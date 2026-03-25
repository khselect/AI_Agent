"""
safety_analytics.py  ─ 단일 파일 버전 (v2.1)
────────────────────────────────────────────────────────────
db_manager + risk_model + UI 를 하나의 파일에 통합
별도 모듈 설치 없이 바로 실행 가능

실행:
    streamlit run safety_analytics.py

필수 패키지:
    pip install streamlit duckdb pandas altair pymupdf4llm \
                langchain-community langchain scikit-learn openpyxl
"""

# ══════════════════════════════════════════════════════════════
# 0. 공통 임포트
# ══════════════════════════════════════════════════════════════
import streamlit as st
import os, sys, json, tempfile, re, io
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from railway_agent.agent_ui import render_agent_tab

import pandas as pd
import numpy as np
import altair as alt

# ── LLM ──────────────────────────────────────────────────────
try:
    from langchain_ollama import ChatOllama
    from langchain.schema import HumanMessage, SystemMessage
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False

# ── PDF ──────────────────────────────────────────────────────
try:
    import pymupdf4llm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# ── DuckDB ───────────────────────────────────────────────────
try:
    import duckdb
    DUCKDB_AVAILABLE = True
except ImportError:
    DUCKDB_AVAILABLE = False
    st.error("duckdb 미설치: `pip install duckdb`")
    st.stop()


# ══════════════════════════════════════════════════════════════
# 1. DB 레이어 (db_manager 인라인)
# ══════════════════════════════════════════════════════════════
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SHARED_DIR = os.path.join(BASE_DIR, "shared")
DB_PATH = os.path.join(SHARED_DIR, "railway_accidents.duckdb")
os.makedirs(SHARED_DIR, exist_ok=True)

DDL_ACCIDENTS = """
CREATE TABLE IF NOT EXISTS accidents (
    id               INTEGER PRIMARY KEY,
    created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    source_file      VARCHAR,
    발생일자          VARCHAR,
    발생시간          VARCHAR,
    등록기관          VARCHAR,
    철도구분          VARCHAR,
    노선              VARCHAR,
    이벤트대분류      VARCHAR,
    이벤트중분류      VARCHAR,
    이벤트소분류      VARCHAR,
    주원인            VARCHAR,
    근본원인그룹      VARCHAR,
    근본원인유형      VARCHAR,
    근본원인상세      VARCHAR,
    직접원인          VARCHAR,
    운행영향유형      VARCHAR,
    지연여부          VARCHAR,
    지연원인          VARCHAR,
    지연원인상세      VARCHAR,
    지연열차수        INTEGER,
    최대지연시간_분   INTEGER,
    총피해인원        INTEGER,
    사망자수          INTEGER,
    부상자수          INTEGER,
    피해액_백만원     DOUBLE,
    행정구역          VARCHAR,
    발생역A           VARCHAR,
    발생역B           VARCHAR,
    장소대분류        VARCHAR,
    장소중분류        VARCHAR,
    상세위치          VARCHAR,
    기상상태          VARCHAR,
    온도              DOUBLE,
    강우량            DOUBLE,
    적설량            DOUBLE,
    대상구분          VARCHAR,
    열차종류          VARCHAR,
    선로유형          VARCHAR,
    신호시스템유형    VARCHAR,
    고장부품명        VARCHAR,
    고장현상          VARCHAR,
    고장원인          VARCHAR,
    조치내용          VARCHAR,
    이벤트개요        VARCHAR,
    데이터출처        VARCHAR,
    risk_score        DOUBLE,
    risk_grade        VARCHAR,
    raw_json          VARCHAR
)
"""

def _get_conn() -> duckdb.DuckDBPyConnection:
    conn = duckdb.connect(DB_PATH)
    conn.execute("CREATE SEQUENCE IF NOT EXISTS accidents_seq START 1")
    conn.execute(DDL_ACCIDENTS)
    return conn

def _safe_int(val, default=0) -> Optional[int]:
    if val is None or str(val).strip() in ("", "null", "NULL", "None", "⬜ 미추출"):
        return default
    try:
        return int(float(str(val)))
    except Exception:
        return default

def _safe_float(val, default=None) -> Optional[float]:
    if val is None or str(val).strip() in ("", "null", "NULL", "None", "⬜ 미추출"):
        return default
    try:
        return float(str(val))
    except Exception:
        return default

def _safe_str(val) -> Optional[str]:
    if val is None or str(val).strip() in ("", "null", "NULL", "None", "⬜ 미추출"):
        return None
    return str(val).strip()

RISK_WEIGHTS = {
    'event_risk': {
        '탈선':85,'충돌':90,'화재':95,'폭발':100,
        '추락':80,'끼임':65,'감전':70,'누출':75,
        '신호무응답':60,'차량고장':40,'궤도틀림':55,
        '전력고장':45,'기타':30,
    },
    'cause_weight': {'인적요인':1.2,'기술적요인':1.0,'환경적요인':0.8},
    'weather_weight': {'맑음':1.0,'흐림':1.1,'비':1.3,'눈':1.5,'안개':1.4,'강풍':1.4},
}

def calculate_risk(extracted: dict) -> tuple:
    """
    위험점수(0~100) 산정
    ─ 인명 : 등가사망지수(EFI) = 사망자 + 부상자 / 100  (철도분야 기준)
    ─ 물적 : 피해액(백만원) / 50, 최대 20점
    ─ 운행 : 최대지연시간_분 / 40, 최대 15점
    ─ 이벤트 유형 가중치 : 최대 15점
    """
    score  = 0.0
    dead   = _safe_int(extracted.get('사망자수'), 0)
    injured= _safe_int(extracted.get('부상자수'), 0)
    damage = _safe_float(extracted.get('피해액(백만원)'), 0) or 0
    delay  = _safe_int(extracted.get('최대지연시간(분)'), 0)
    evt    = str(extracted.get('이벤트대분류','') or '')
    sub    = str(extracted.get('이벤트소분류','') or '')

    # 등가사망지수 (EFI): 부상자 100명 = 사망자 1명
    efi = dead + injured / 100.0
    score += min(efi * 20, 40)          # 인명 최대 40점
    score += min(damage / 50, 20)       # 물적 최대 20점
    score += min(delay / 40, 15)        # 지연 최대 15점

    high_kw = ['탈선','충돌','화재','폭발','추락','붕괴']
    if any(k in sub for k in high_kw): score += 15
    elif evt == '사고': score += 10
    elif evt == '장애': score += 5

    # ── 인명피해 절대 최솟값 (철도안전 기준) ──────────────────
    # 사망자 1명 이상: 반드시 High 이상 (score ≥ 60)
    if dead >= 1:
        score = max(score, 60.0)
    # 사망자 3명 이상 또는 부상자 20명 이상: Critical 권고 (score ≥ 80)
    if dead >= 3 or injured >= 20:
        score = max(score, 80.0)
    # 사망자 5명 이상: Critical 최상위 (score ≥ 90)
    if dead >= 5:
        score = max(score, 90.0)

    score = min(round(score, 1), 100)
    grade = 'High' if score >= 60 else ('Medium' if score >= 25 else 'Low')
    return score, grade

def insert_accident(extracted: dict, source_file: str = "") -> int:
    conn = _get_conn()
    risk_score, risk_grade = calculate_risk(extracted)
    row_id = conn.execute("SELECT nextval('accidents_seq')").fetchone()[0]
    conn.execute("""
        INSERT INTO accidents VALUES (
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
        )
    """, [
        row_id, datetime.now(), source_file,
        _safe_str(extracted.get('발생일자')), _safe_str(extracted.get('발생시간')),
        _safe_str(extracted.get('등록기관')), _safe_str(extracted.get('철도구분')),
        _safe_str(extracted.get('노선')),
        _safe_str(extracted.get('이벤트대분류')), _safe_str(extracted.get('이벤트중분류')),
        _safe_str(extracted.get('이벤트소분류')), _safe_str(extracted.get('주원인')),
        _safe_str(extracted.get('근본원인그룹')), _safe_str(extracted.get('근본원인유형')),
        _safe_str(extracted.get('근본원인상세')), _safe_str(extracted.get('직접원인')),
        _safe_str(extracted.get('운행영향유형')), _safe_str(extracted.get('지연여부')),
        _safe_str(extracted.get('지연원인')), _safe_str(extracted.get('지연원인상세')),
        _safe_int(extracted.get('지연열차수')), _safe_int(extracted.get('최대지연시간(분)')),
        _safe_int(extracted.get('총피해인원')), _safe_int(extracted.get('사망자수')),
        _safe_int(extracted.get('부상자수')), _safe_float(extracted.get('피해액(백만원)')),
        _safe_str(extracted.get('행정구역')), _safe_str(extracted.get('발생역A')),
        _safe_str(extracted.get('발생역B')), _safe_str(extracted.get('장소대분류')),
        _safe_str(extracted.get('장소중분류')), _safe_str(extracted.get('상세위치')),
        _safe_str(extracted.get('기상상태')), _safe_float(extracted.get('온도')),
        _safe_float(extracted.get('강우량')), _safe_float(extracted.get('적설량')),
        _safe_str(extracted.get('대상구분')), _safe_str(extracted.get('열차종류')),
        _safe_str(extracted.get('선로유형')), _safe_str(extracted.get('신호시스템유형')),
        _safe_str(extracted.get('고장부품명')), _safe_str(extracted.get('고장현상')),
        _safe_str(extracted.get('고장원인')), _safe_str(extracted.get('조치내용')),
        _safe_str(extracted.get('이벤트개요')), _safe_str(extracted.get('데이터 출처')),
        risk_score, risk_grade,
        json.dumps(extracted, ensure_ascii=False),
    ])
    conn.close()
    return row_id

def get_all_accidents() -> pd.DataFrame:
    conn = _get_conn()
    df = conn.execute("SELECT * FROM accidents ORDER BY id DESC").df()
    conn.close()
    return df

def get_accident_count() -> int:
    conn = _get_conn()
    n = conn.execute("SELECT COUNT(*) FROM accidents").fetchone()[0]
    conn.close()
    return n

def delete_accident(row_id: int):
    conn = _get_conn()
    conn.execute("DELETE FROM accidents WHERE id = ?", [row_id])
    conn.close()


# ══════════════════════════════════════════════════════════════
# 2. 예측 모델 (risk_model 인라인)
# ══════════════════════════════════════════════════════════════

def find_similar_accidents(df: pd.DataFrame, query: dict, top_k: int = 5) -> pd.DataFrame:
    if df.empty or len(df) < 2:
        return df.head(top_k)
    scores = pd.Series(0.0, index=df.index)
    for col, weight in [('노선',3.0),('이벤트소분류',4.0),('근본원인그룹',2.0),
                         ('기상상태',1.0),('열차종류',1.5),('장소대분류',2.0)]:
        if col in df.columns and col in query and query[col]:
            scores += df[col].eq(query[col]).astype(float) * weight
    df = df.copy()
    df['_sim'] = scores
    return df.nlargest(top_k, '_sim').drop(columns=['_sim'])

def predict_risk_statistical(df: pd.DataFrame, scenario: dict) -> dict:
    if df.empty:
        return {'predicted_score':50.0,'predicted_grade':'Medium',
                'confidence':'낮음 (데이터 없음)','basis':'데이터 없음','similar_count':0}
    similar = find_similar_accidents(df, scenario, top_k=20)
    n = len(similar)
    base_score = similar['risk_score'].mean() if 'risk_score' in similar.columns and similar['risk_score'].notna().any() else 50.0
    weather = scenario.get('기상상태','맑음')
    w_mult  = RISK_WEIGHTS['weather_weight'].get(weather, 1.0)
    cause   = scenario.get('근본원인그룹','')
    c_mult  = RISK_WEIGHTS['cause_weight'].get(cause, 1.0)
    evt_sub = scenario.get('이벤트소분류','')
    evt_base= RISK_WEIGHTS['event_risk'].get(evt_sub, 0)
    final   = (base_score*0.7 + evt_base*0.3)*w_mult*c_mult if evt_base > 0 else base_score*w_mult*c_mult
    final   = min(round(final,1), 100)
    grade   = 'High' if final>=60 else ('Medium' if final>=25 else 'Low')
    conf    = '높음' if n>=10 else ('보통' if n>=5 else '낮음 (유사 사례 부족)')
    basis   = f"유사 {n}건 평균 {base_score:.0f}점 / 기상({weather}) ×{w_mult} / 원인({cause or '미상'}) ×{c_mult:.1f}"
    return {'predicted_score':final,'predicted_grade':grade,'confidence':conf,
            'basis':basis,'similar_count':n,'similar_df':similar}

def run_anomaly_detection(df: pd.DataFrame) -> pd.DataFrame:
    if len(df) < 10:
        return df.assign(anomaly_score=None, is_anomaly=False)
    try:
        from sklearn.ensemble import IsolationForest
        from sklearn.preprocessing import LabelEncoder
        import warnings; warnings.filterwarnings('ignore')
        nums = ['사망자수','부상자수','피해액_백만원','최대지연시간_분','지연열차수','risk_score']
        cats = ['이벤트소분류','근본원인그룹','기상상태']
        work = df.copy()
        for c in nums:
            work[c] = pd.to_numeric(work.get(c, pd.Series([0]*len(work))), errors='coerce').fillna(0)
        for c in cats:
            if c in work.columns:
                le = LabelEncoder()
                work[c+'_enc'] = le.fit_transform(work[c].fillna('unknown'))
        X_cols = nums + [c+'_enc' for c in cats if c in work.columns]
        X = work[[c for c in X_cols if c in work.columns]].values
        model = IsolationForest(contamination=0.1, random_state=42)
        sc = model.fit_predict(X)
        df = df.copy()
        df['anomaly_score'] = np.round(-model.decision_function(X)*100, 1)
        df['is_anomaly'] = sc == -1
        return df
    except Exception:
        return df.assign(anomaly_score=None, is_anomaly=False)

SCENARIO_TEMPLATES = {
    ('탈선','인적요인'):["신호 무시/오인 → 분기기 통과 중 탈선","과속 운행 → 곡선 구간 탈선","운전 부주의 → 차량 기지 내 탈선"],
    ('탈선','기술적요인'):["궤도 틀림 누적 → 열차 동요 → 탈선","차축/차륜 결함 → 선로 이탈","분기기 전환 미완료 → 탈선"],
    ('충돌','인적요인'):["정지신호 현시 구간 과주 → 후행 열차 추돌","관제 통신 오류 → 동일 선로 대향 진입"],
    ('화재','기술적요인'):["전동차 전장품 과부하 → 발화 → 객실 확산","제동장치 과열 → 차체 하부 발화"],
}
MITIGATIONS = {
    ('탈선','인적요인'):"승무원 신호 준수 교육 강화, ATP 점검, 피로 관리 절차 수립",
    ('탈선','기술적요인'):"궤도 정기 검측 주기 단축, 차륜/차축 비파괴 검사, 분기기 센서 이중화",
    ('충돌','인적요인'):"관제 통신 프로토콜 재정비, CTC 경보 강화, 열차 방호 장치 설치",
    ('화재','기술적요인'):"전장품 내열 등급 상향, 자동 소화 시스템, 화재 감지 센서 추가",
}

def generate_scenarios(event_type, cause_group, line="", weather="맑음") -> list:
    templates = SCENARIO_TEMPLATES.get((event_type, cause_group), [
        f"{event_type} 발생 → 현장 대응 지연 → 2차 피해 확대",
        f"초기 {event_type} 징후 미인지 → 적시 조치 실패 → 사고 심화",
    ])
    wf = {'눈':'적설로 제동거리 증가','비':'우천으로 시야 제한','안개':'안개로 신호 확인 지연'}.get(weather,'')
    scenarios = []
    for i, tmpl in enumerate(templates):
        desc = (f"[{line}] " if line else "") + tmpl + (f" + {wf}" if wf else "")
        sev = 'High' if any(k in desc for k in ['추돌','화재','감전','폭발']) else ('Medium' if i==0 else 'Low')
        scenarios.append({'no':i+1,'scenario':desc,'severity':sev,
                          'mitigation':MITIGATIONS.get((event_type,cause_group),
                                       "정기 안전 점검 강화, 위험 요소 모니터링, 비상 대응 훈련")})
    return scenarios

def analyze_trends(df: pd.DataFrame) -> dict:
    if df.empty: return {}
    return {
        'total': len(df),
        'high_risk': int((df['risk_grade']=='High').sum()) if 'risk_grade' in df.columns else 0,
        'avg_risk_score': float(df['risk_score'].mean()) if 'risk_score' in df.columns else 0,
        'total_deaths': int(df['사망자수'].fillna(0).sum()) if '사망자수' in df.columns else 0,
        'total_injured': int(df['부상자수'].fillna(0).sum()) if '부상자수' in df.columns else 0,
    }


# ══════════════════════════════════════════════════════════════
# 3. PDF 추출 (report_extractor_v2 로직 인라인)
# ══════════════════════════════════════════════════════════════
COLUMNS = [
    ("발생일자","이벤트 발생 날짜. YYYY-MM-DD"),
    ("발생시간","이벤트 발생 시간. HH:MM"),
    ("등록기관","데이터를 등록·보고한 기관명"),
    ("철도구분","일반철도/도시철도/고속철도"),
    ("노선","노선명"),
    ("이벤트대분류","사고/장애/고장"),
    ("이벤트중분류","차량/신호/선로/전력/외부요인 등"),
    ("이벤트소분류","탈선, 충돌, 화재 등"),
    ("주원인","1차 원인 요약"),
    ("근본원인그룹","인적요인/기술적요인/환경적요인"),
    ("근본원인유형","운전취급, 열차차량설비 등"),
    ("근본원인상세","상세 원인 설명"),
    ("직접원인","직접 원인"),
    ("운행영향유형","운행중단/지연운행/서행운전"),
    ("지연여부","지연/무지연"),
    ("지연원인","지연 주요 원인"),
    ("지연원인상세","지연 상세 사유"),
    ("지연열차수","숫자"),
    ("최대지연시간(분)","숫자"),
    ("총피해인원","숫자"),
    ("사망자수","숫자"),
    ("부상자수","숫자"),
    ("피해액(백만원)","숫자"),
    ("행정구역","행정 주소"),
    ("발생역A","기준역"),
    ("발생역B","인접역"),
    ("장소대분류","역/본선/기지"),
    ("장소중분류","구내선로/본선/승강장"),
    ("상세위치","상세 위치"),
    ("기상상태","맑음/흐림/비/눈/안개"),
    ("온도","℃ 숫자"),
    ("강우량","mm 숫자"),
    ("적설량","cm 숫자"),
    ("대상구분","열차/차량/설비"),
    ("열차종류","전동열차/화물열차/여객열차/KTX"),
    ("선로유형","지상/지하/교량"),
    ("신호시스템유형","ATP/ATO, 자동폐색 등"),
    ("고장부품명","부품명"),
    ("고장현상","현상 설명"),
    ("고장원인","기술적 원인"),
    ("조치내용","조치 내용 요약"),
    ("이벤트개요","3~5문장 요약"),
    ("데이터 출처","출처"),
]
COLUMN_NAMES = [c[0] for c in COLUMNS]

BATCHES = [COLUMNS[0:9], COLUMNS[9:18], COLUMNS[18:26], COLUMNS[26:34], COLUMNS[34:]]
BATCH_NAMES = ["기본정보","원인·지연","피해·위치A","위치·기상","선로·고장·개요"]

def _is_qwen3(m: str) -> bool:
    return "qwen3" in m.lower()

def _clean_llm(raw: str) -> str:
    """LLM 응답에서 노이즈 제거 — qwen3 think 블록·마크다운 코드펜스 등"""
    # 1) 완결된 <think>...</think> 제거
    raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL)
    # 2) 미완결 <think> (닫힘 태그 없음) → <think> 이후 첫 { 전까지 제거
    if "<think>" in raw:
        brace = raw.find("{", raw.find("<think>"))
        if brace != -1:
            raw = raw[brace:]
        else:
            raw = re.sub(r"<think>.*", "", raw, flags=re.DOTALL)
    # 3) 마크다운 코드블록 제거
    raw = re.sub(r"```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
    raw = raw.replace("```", "")
    # 4) JSON 앞 자연어 서문 제거 (첫 { 이전 텍스트)
    brace = raw.find("{")
    if brace > 0:
        raw = raw[brace:]
    return raw.strip()

def _safe_json(text: str) -> dict:
    """강화된 JSON 파싱 — Python 3.9 호환, 6단계 fallback"""
    text = _clean_llm(text)

    def _repair(s):
        s = re.sub(r',\s*([}\]])', r'\1', s)           # trailing comma
        s = re.sub(r'//[^\n]*', '', s)                  # // 주석
        s = re.sub(r'/\*.*?\*/', '', s, flags=re.DOTALL)
        return s

    def _fix_sq(s):
        """단따옴표 키/값만 쌍따옴표로 교체 (내부 아포스트로피 보호)"""
        s = re.sub(r"'([^'\n]{1,80})'\s*:", r'"\1":', s)
        s = re.sub(r":\s*'([^'\n]*?)'", r': "\1"', s)
        return s

    blk = (re.search(r'\{[\s\S]*\}', text) or None)
    blk_str = blk.group() if blk else ""

    for candidate in ([text, blk_str] if blk_str else [text]):
        for transform in [lambda s: s, _repair, _fix_sq,
                          lambda s: _repair(_fix_sq(s))]:
            try:
                t = transform(candidate)
                r = json.loads(t)
                if isinstance(r, dict) and r:
                    return r
            except Exception:
                pass

    # 최후 수단: 키-값 정규식 스캔
    result = {}
    for key, _ in COLUMNS:
        k = re.escape(key)
        ms = re.search(r'["\']?' + k + r'["\']?\s*:\s*["\']([^"\'\\n]{0,300})["\']', text)
        if ms and ms.group(1).strip() not in ('null', 'NULL', 'None', ''):
            result[key] = ms.group(1).strip()
            continue
        mn = re.search(r'["\']?' + k + r'["\']?\s*:\s*(-?\d+\.?\d*)', text)
        if mn:
            result[key] = mn.group(1)
    return result

def _regex_base(t: str) -> dict:
    """정규식 기반 1차 추출 — LLM 미사용 시 또는 LLM 실패 필드 보완"""
    d = {}
    # ── 날짜·시간 ─────────────────────────────────────────────
    dm = re.search(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', t)
    if dm: d['발생일자'] = f"{dm.group(1)}-{int(dm.group(2)):02d}-{int(dm.group(3)):02d}"
    tm = re.search(r'(\d{1,2})시\s*(\d{2})분', t)
    if tm: d['발생시간'] = f"{int(tm.group(1)):02d}:{tm.group(2)}"

    # ── 기관·철도구분 ─────────────────────────────────────────
    AGENCIES = ['서울교통공사','KORAIL','한국철도공사','부산교통공사','대구도시철도',
                '광주도시철도','대전도시철도','인천교통공사','SR','공항철도']
    for ag in AGENCIES:
        if ag in t: d['등록기관'] = ag; break
    if 'KTX' in t or '고속철도' in t or 'SRT' in t: d['철도구분'] = '고속철도'
    elif any(k in t for k in ['호선','지하철','도시철도']): d['철도구분'] = '도시철도'
    else: d['철도구분'] = '일반철도'

    # ── 노선 ──────────────────────────────────────────────────
    nm = re.search(
        r'(서울\s*\d+호선|부산\s*\d+호선|대구\s*\d+호선|인천\s*\d+호선|'
        r'광주\s*\d+호선|대전\s*\d+호선|경부선|경인선|수인선|중앙선|'
        r'분당선|신분당선|공항철도|경강선|KTX|SRT)', t
    )
    if nm: d['노선'] = nm.group(1).replace(' ', '')

    # ── 이벤트 분류 ───────────────────────────────────────────
    EVT_MAP = {
        '탈선': ('사고','차량','탈선'), '충돌': ('사고','차량','충돌'),
        '화재': ('사고','차량','화재'), '추락': ('사고','인적','추락'),
        '신호장애': ('장애','신호','신호장애'), '전력장애': ('장애','전력','전력장애'),
        '차량고장': ('장애','차량','차량고장'), '선로장애': ('장애','선로','선로장애'),
    }
    for kw, (大, 中, 小) in EVT_MAP.items():
        if kw in t:
            d.update({'이벤트대분류':大, '이벤트중분류':中, '이벤트소분류':小})
            break

    # ── 인명피해 ──────────────────────────────────────────────
    dead = re.search(r'사망자?\s*(\d+)\s*명', t)
    d['사망자수'] = dead.group(1) if dead else '0'
    inj = re.search(r'부상자?\s*(\d+)\s*명', t)
    d['부상자수'] = inj.group(1) if inj else '0'
    d['총피해인원'] = str(int(d.get('사망자수','0') or 0) + int(d.get('부상자수','0') or 0))

    # ── 피해액 ────────────────────────────────────────────────
    dmg = re.search(r'(?:총\s*)?([\d,]+)\s*백만\s*원', t)
    if dmg: d['피해액(백만원)'] = dmg.group(1).replace(',','')
    else:
        dmg2 = re.search(r'([\d,]+)\s*원(?!권)', t)
        if dmg2:
            won = int(dmg2.group(1).replace(',',''))
            if won >= 1_000_000:
                d['피해액(백만원)'] = str(round(won / 1_000_000, 1))

    # ── 지연 ──────────────────────────────────────────────────
    delay = re.search(r'(\d+)\s*분(?:\s*(?:지연|운휴|중단))', t)
    if delay: d['최대지연시간(분)'] = delay.group(1)
    dly_cnt = re.search(r'(\d+)\s*(?:개|편)?\s*열차(?:\s*지연)?', t)
    if dly_cnt: d['지연열차수'] = dly_cnt.group(1)
    if any(k in t for k in ['운행 중단','운행중단','운휴']): d['지연여부'] = '지연'
    elif any(k in t for k in ['지연','서행']): d['지연여부'] = '지연'
    else: d['지연여부'] = '무지연'

    # ── 위치 ──────────────────────────────────────────────────
    sta = re.search(r'([가-힣]+역)(?:\s*(\d+)번\s*승강장)?', t)
    if sta:
        d['발생역A'] = sta.group(1)
        if sta.group(2): d['상세위치'] = f"{sta.group(2)}번 승강장"
    if '승강장' in t: d.setdefault('장소중분류', '승강장')
    if '구내선로' in t: d.setdefault('장소중분류', '구내선로')
    if '역' in t: d.setdefault('장소대분류', '역')
    elif '기지' in t or '차량기지' in t: d['장소대분류'] = '기지'

    # ── 기상·환경 ─────────────────────────────────────────────
    for kw, wv in {'맑았':'맑음','맑음':'맑음','흐림':'흐림','비':'비','눈':'눈','안개':'안개'}.items():
        if kw in t: d['기상상태'] = wv; break
    temp = re.search(r'(-?\d+(?:\.\d+)?)\s*℃', t)
    if temp: d['온도'] = temp.group(1)

    # ── 기술 ──────────────────────────────────────────────────
    if '지하' in t: d['선로유형'] = '지하'
    elif '교량' in t: d['선로유형'] = '교량'
    else: d.setdefault('선로유형', '지상')
    if 'ATP' in t and 'ATO' in t: d['신호시스템유형'] = 'ATP/ATO'
    elif 'ATP' in t: d['신호시스템유형'] = 'ATP'
    elif '자동폐색' in t: d['신호시스템유형'] = '자동폐색'

    # ── 열차종류 ──────────────────────────────────────────────
    if '전동열차' in t or '전동차' in t: d['열차종류'] = '전동열차'
    elif 'KTX' in t: d['열차종류'] = 'KTX'
    elif 'SRT' in t: d['열차종류'] = 'SRT'
    elif '화물' in t: d['열차종류'] = '화물열차'

    return d

# 배치별 보고서 텍스트 슬라이스 전략
# - 기본정보(배치0): 앞부분 집중 (제목·일시·노선)
# - 원인·지연(배치1): 중간부분 (조사결과·원인분석)
# - 피해·위치(배치2): 앞+중간 (피해현황·사고위치)
# - 위치·기상(배치3): 중간+뒷부분 (현장조건·기상)
# - 선로·고장·개요(배치4): 뒷부분 전체 (기술분석·조치)
BATCH_SLICE = [
    (0,    10000),   # 배치0 기본정보: 앞 10000자
    (3000, 16000),   # 배치1 원인·지연: 3000~16000자
    (0,    12000),   # 배치2 피해·위치A: 앞 12000자
    (5000, 18000),   # 배치3 위치·기상: 5000~18000자
    (8000, None),    # 배치4 선로·고장·개요: 8000자 이후 전체
]

def _slice_text(report_text: str, batch_idx: int) -> str:
    """배치 인덱스에 따라 보고서 적절 구간 추출. 최대 12000자."""
    start, end = BATCH_SLICE[batch_idx]
    chunk = report_text[start:end] if end else report_text[start:]
    return chunk[:12000]  # 안전 상한 (num_ctx=32768 기준)

def _build_batch_prompt(batch_cols, report_text, model_name, batch_idx=0):
    prefix = "/no_think\n" if _is_qwen3(model_name) else ""
    schema_keys = ", ".join(f'"{n}"' for n, _ in batch_cols)
    guide = "\n".join(f'  - "{n}": {desc}' for n, desc in batch_cols)
    text_chunk = _slice_text(report_text, batch_idx)

    # Few-shot 예시 (숫자 필드 명확화)
    num_fields = {n for n, d in batch_cols if any(k in d for k in ['숫자','수','분','℃','mm','cm'])}
    num_note = ""
    if num_fields:
        examples = ", ".join(f'"{n}": 0' for n in list(num_fields)[:3])
        num_note = f'\n숫자 필드는 따옴표 없이 숫자로만. 예: {{{examples}}}\n'

    # 출력 템플릿: 키=null 형태로 LLM이 구조 그대로 채우도록
    json_template = "{" + ", ".join(f'"{n}": null' for n, _ in batch_cols) + "}"

    return f"""{prefix}You are a railway accident report data extractor.
Extract ONLY the fields listed below from the [REPORT] and return a single JSON object.

STRICT RULES:
1. Output ONLY the JSON object — no explanation, no markdown, no code blocks
2. Use null for missing or unclear fields
3. Date format: "YYYY-MM-DD", Time format: "HH:MM"
4. Numeric fields: use numbers without quotes (e.g. 3, not "3"){num_note}
FIELDS TO EXTRACT:
{guide}

OUTPUT TEMPLATE (fill in the values, keep null if not found):
{json_template}

[REPORT]
{text_chunk}

JSON:"""

def extract_from_pdf(pdf_bytes: bytes, model_name: str, progress_fn=None) -> tuple:
    if not PDF_AVAILABLE:
        raise RuntimeError("pymupdf4llm 미설치: pip install pymupdf4llm")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes); tmp_path = tmp.name

    try:
        if progress_fn: progress_fn(0.05, "📖 PDF 텍스트 추출 중...")
        report_text = pymupdf4llm.to_markdown(tmp_path)
        result = _regex_base(report_text)

        if LLM_AVAILABLE:
            llm = ChatOllama(
                model=model_name, base_url="http://127.0.0.1:11434",
                temperature=0,
                num_ctx=32768,   # ↑ 8192→32768: 한국어 장문 보고서 전체 처리
                num_predict=2048,
            )
            sys_msg = SystemMessage(content=(
                "You are a structured data extractor. "
                "Output ONLY a valid JSON object. "
                "No markdown, no code blocks, no explanations."
            ))
            for i, batch in enumerate(BATCHES):
                pct = 0.15 + 0.65 * i / len(BATCHES)
                if progress_fn: progress_fn(pct, f"🤖 배치 {i+1}/{len(BATCHES)}: {BATCH_NAMES[i]} 추출 중...")
                try:
                    # batch_idx 전달 → 배치별 최적 텍스트 구간 사용
                    prompt = _build_batch_prompt(batch, report_text, model_name, batch_idx=i)
                    resp = llm.invoke([sys_msg, HumanMessage(content=prompt)])
                    batch_result = _safe_json(resp.content)
                    for col_name, _ in batch:
                        val = batch_result.get(col_name)
                        if val is not None and str(val).strip() not in ("","null","NULL","None",""):
                            result[col_name] = str(val).strip()
                except Exception as e:
                    if progress_fn: progress_fn(pct, f"⚠️ 배치 {i+1} 오류: {e}")

        result['데이터 출처'] = result.get('데이터 출처') or 'PDF 자동 추출'
        # 추출률 계산 및 로깅
        total_fields = len(COLUMN_NAMES)
        extracted_fields = sum(
            1 for k in COLUMN_NAMES
            if result.get(k) and str(result[k]).strip() not in ('','None','null','NULL')
        )
        rate = extracted_fields / total_fields * 100
        msg = f"✅ 추출 완료 ({extracted_fields}/{total_fields}개 필드, {rate:.0f}%)"
        if progress_fn: progress_fn(0.95, msg)
        return result, report_text
    finally:
        if os.path.exists(tmp_path): os.remove(tmp_path)


# ══════════════════════════════════════════════════════════════
# 4. Streamlit UI
# ══════════════════════════════════════════════════════════════
st.set_page_config(page_title="🚄 철도 사고 분석 시스템", layout="wide", initial_sidebar_state="expanded")
st.title("🚄 철도 사고조사 데이터 분석 시스템")

# ── 사이드바 ──────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 설정")
    CONFIG_FILE = os.path.join(SHARED_DIR, "system_config.json")
    default_model = "qwen2.5:7b-instruct"
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                default_model = json.load(f).get("selected_model", default_model)
        except Exception:
            pass

    MODELS = ["qwen3:8b","qwen2.5:7b-instruct","llama3.1:8b"]
    try: midx = MODELS.index(default_model)
    except ValueError: midx = 1

    model_name = st.selectbox("🤖 LLM 모델", MODELS, index=midx)
    if _is_qwen3(model_name):
        st.info("💡 qwen3: /no_think 자동 적용")

    st.divider()
    total_records = get_accident_count()
    st.metric("누적 사고 데이터", f"{total_records}건")

    phase = "Phase 3 🟢" if total_records >= 200 else ("Phase 2 🟡" if total_records >= 50 else "Phase 1 🔴")
    st.caption(f"예측 모델: {phase}")
    st.caption(f"DB: shared/railway_accidents.duckdb")

    with st.expander("📌 Phase 안내"):
        st.markdown("""
- **Phase 1** (0~49건): 규칙+통계 기반
- **Phase 2** (50건~): Isolation Forest 이상탐지
- **Phase 3** (200건~): Random Forest 분류
        """)

# ── 탭 ───────────────────────────────────────────────────────
tab1, tab_data, tab2, tab3, tab4, tab_agent = st.tabs([
    "📥 보고서 입력", "📋 데이터 조회/관리", "📊 대시보드", "⚠️ 위험도 평가", "🔮 위험 예측", "🤖 AI 에이전트"
])


# ════════════════════════════════════════
# TAB 1. 보고서 입력
# ════════════════════════════════════════
with tab1:
    st.subheader("📥 사고조사보고서 입력")
    input_mode = st.radio("입력 방식", ["📄 PDF 자동 추출", "✏️ 수동 직접 입력"], horizontal=True)
    st.divider()

    if input_mode == "📄 PDF 자동 추출":
        if not PDF_AVAILABLE:
            st.error("pymupdf4llm 미설치: `pip install pymupdf4llm`")
        else:
            uploaded = st.file_uploader("PDF 보고서 업로드", type=["pdf"])
            if uploaded:
                st.success(f"✅ {uploaded.name} ({uploaded.size/1024:.1f} KB)")
                if st.button("🚀 추출 + DB 저장", type="primary"):
                    prog = st.progress(0.0); stat = st.empty()
                    def upd(pct, msg): prog.progress(pct); stat.info(msg)
                    try:
                        extracted, _ = extract_from_pdf(uploaded.getvalue(), model_name, upd)
                        row_id = insert_accident(extracted, uploaded.name)
                        prog.progress(1.0); stat.success(f"🎉 저장 완료! (DB ID: {row_id})")
                        score, grade = calculate_risk(extracted)
                        filled = sum(1 for k in COLUMN_NAMES if extracted.get(k) not in (None,'','null','NULL'))
                        # ★ 탭 전환 후에도 결과 유지되도록 session_state에 저장
                        st.session_state["tab1_result"] = {
                            "row_id":    row_id,
                            "source":    uploaded.name,
                            "grade":     grade,
                            "score":     score,
                            "filled":    filled,
                            "extracted": extracted,
                        }
                    except Exception as e:
                        import traceback; st.error(f"오류: {e}"); st.text(traceback.format_exc())

            # ── 추출 결과 표시 (session_state에서 복원 → 탭 전환 무관) ──
            if "tab1_result" in st.session_state:
                _r = st.session_state["tab1_result"]
                st.success(f"🎉 저장 완료! (DB ID: {_r['row_id']}) · 파일: {_r['source']}")
                c1, c2, c3 = st.columns(3)
                c1.metric("위험 등급",  _r["grade"])
                c2.metric("위험 점수",  f"{_r['score']}점")
                c3.metric("추출 필드",  f"{_r['filled']}/{len(COLUMN_NAMES)}")
                with st.expander("📋 추출 결과 (노란색=미추출)", expanded=True):
                    _EMPTY = ('', None, 'null', 'NULL', 'None')
                    _rows = []
                    for n, desc in COLUMNS:
                        val = _r["extracted"].get(n, "")
                        is_empty = (val is None or str(val).strip() in _EMPTY)
                        _rows.append({
                            "필드명": n,
                            "추출값": "⬜ 미추출" if is_empty else str(val),
                            "설명":   desc,
                        })
                    _df_result = pd.DataFrame(_rows)
                    def _hl(row):
                        if row["추출값"] == "⬜ 미추출":
                            return ["background-color:#FFF9C4"] * len(row)
                        return [""] * len(row)
                    styled = _df_result.style.apply(_hl, axis=1)
                    st.dataframe(styled, use_container_width=True,
                                 hide_index=True, height=520)
                    _batch_stats = []
                    for bi, batch in enumerate(BATCHES):
                        _filled_b = sum(
                            1 for n, _ in batch
                            if _r["extracted"].get(n)
                            and str(_r["extracted"].get(n)).strip() not in _EMPTY
                        )
                        _batch_stats.append(f"배치{bi+1}({BATCH_NAMES[bi]}): {_filled_b}/{len(batch)}")
                    st.caption("  |  ".join(_batch_stats))
    else:
        manual = {}
        with st.form("manual_form"):
            c1, c2 = st.columns(2)
            with c1:
                manual['발생일자']     = st.text_input("발생일자 (YYYY-MM-DD)")
                manual['발생시간']     = st.text_input("발생시간 (HH:MM)")
                manual['등록기관']     = st.text_input("등록기관")
                manual['철도구분']     = st.selectbox("철도구분", ["도시철도","일반철도","고속철도"])
                manual['노선']         = st.text_input("노선")
                manual['이벤트대분류'] = st.selectbox("이벤트대분류", ["사고","장애","고장"])
                manual['이벤트중분류'] = st.text_input("이벤트중분류")
                manual['이벤트소분류'] = st.text_input("이벤트소분류")
            with c2:
                manual['근본원인그룹'] = st.selectbox("근본원인그룹", ["인적요인","기술적요인","환경적요인"])
                manual['주원인']       = st.text_input("주원인")
                manual['사망자수']     = st.number_input("사망자수", 0, 100, 0)
                manual['부상자수']     = st.number_input("부상자수", 0, 999, 0)
                manual['피해액(백만원)']     = st.number_input("피해액(백만원)", 0.0, value=0.0)
                manual['최대지연시간(분)'] = st.number_input("최대지연시간(분)", 0, value=0)
                manual['발생역A']      = st.text_input("발생역A")
                manual['기상상태']     = st.selectbox("기상상태", ["맑음","흐림","비","눈","안개"])
            manual['이벤트개요']   = st.text_area("이벤트 개요", height=80)
            manual['조치내용']     = st.text_area("조치내용", height=60)
            manual['데이터 출처']  = st.text_input("데이터 출처", "수동 입력")
            if st.form_submit_button("💾 DB 저장", type="primary", use_container_width=True):
                for k in ['사망자수','부상자수','피해액(백만원)','최대지연시간(분)']:
                    manual[k] = str(manual[k])
                row_id = insert_accident(manual, "수동입력")
                score, grade = calculate_risk(manual)
                st.success(f"✅ 저장 완료 (ID: {row_id}) | 위험등급: {grade} ({score}점)")
                st.rerun()


# ════════════════════════════════════════
# TAB DATA. 데이터 조회 / 엑셀 다운·업로드
# ════════════════════════════════════════

# DB 컬럼명 ↔ 43개 필드명 매핑
_DB_TO_FIELD = {
    "발생일자":        "발생일자",
    "발생시간":        "발생시간",
    "등록기관":        "등록기관",
    "철도구분":        "철도구분",
    "노선":            "노선",
    "이벤트대분류":    "이벤트대분류",
    "이벤트중분류":    "이벤트중분류",
    "이벤트소분류":    "이벤트소분류",
    "주원인":          "주원인",
    "근본원인그룹":    "근본원인그룹",
    "근본원인유형":    "근본원인유형",
    "근본원인상세":    "근본원인상세",
    "직접원인":        "직접원인",
    "운행영향유형":    "운행영향유형",
    "지연여부":        "지연여부",
    "지연원인":        "지연원인",
    "지연원인상세":    "지연원인상세",
    "지연열차수":      "지연열차수",
    "최대지연시간_분": "최대지연시간(분)",
    "총피해인원":      "총피해인원",
    "사망자수":        "사망자수",
    "부상자수":        "부상자수",
    "피해액_백만원":   "피해액(백만원)",
    "행정구역":        "행정구역",
    "발생역A":         "발생역A",
    "발생역B":         "발생역B",
    "장소대분류":      "장소대분류",
    "장소중분류":      "장소중분류",
    "상세위치":        "상세위치",
    "기상상태":        "기상상태",
    "온도":            "온도",
    "강우량":          "강우량",
    "적설량":          "적설량",
    "대상구분":        "대상구분",
    "열차종류":        "열차종류",
    "선로유형":        "선로유형",
    "신호시스템유형":  "신호시스템유형",
    "고장부품명":      "고장부품명",
    "고장현상":        "고장현상",
    "고장원인":        "고장원인",
    "조치내용":        "조치내용",
    "이벤트개요":      "이벤트개요",
    "데이터출처":      "데이터 출처",
}
_FIELD_COLS   = [v for v in _DB_TO_FIELD.values()]   # 43개 필드명 순서
_EXTRA_COLS   = ["id", "risk_score", "risk_grade"]    # 부가 컬럼


def _df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """DataFrame → xlsx bytes (openpyxl 스타일 적용)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "사고데이터"

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="EEF4FB")

    GRADE_FILLS = {
        "Critical": PatternFill("solid", start_color="FADBD8"),
        "High":     PatternFill("solid", start_color="FDEBD0"),
        "Medium":   PatternFill("solid", start_color="FEF9E7"),
        "Low":      PatternFill("solid", start_color="EAFAF1"),
    }

    headers = list(df.columns)
    ws.row_dimensions[1].height = 24
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = border
        # 컬럼 너비 자동 조정
        ws.column_dimensions[c.column_letter].width = max(12, min(len(str(h)) * 2, 40))

    grade_col_idx = headers.index("risk_grade") + 1 if "risk_grade" in headers else None

    for ri, row in enumerate(df.itertuples(index=False), 2):
        grade = str(getattr(row, "risk_grade", "")) if grade_col_idx else ""
        row_fill = GRADE_FILLS.get(grade, alt_fill if ri % 2 == 0 else None)
        for ci, val in enumerate(row, 1):
            cell_val = "" if val is None or (isinstance(val, float) and np.isnan(val)) else val
            c = ws.cell(row=ri, column=ci, value=cell_val)
            c.font      = Font(name="Arial", size=9)
            c.alignment = left
            c.border    = border
            if row_fill:
                c.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_download_df(df_raw: pd.DataFrame) -> pd.DataFrame:
    """DB 조회 결과 → 43개 필드 + 부가컬럼 DataFrame"""
    rows = []
    for _, r in df_raw.iterrows():
        row = {}
        for db_col, field_col in _DB_TO_FIELD.items():
            row[field_col] = r.get(db_col, "")
        for ec in _EXTRA_COLS:
            row[ec] = r.get(ec, "")
        rows.append(row)
    cols_order = ["id"] + _FIELD_COLS + ["risk_score", "risk_grade"]
    df_out = pd.DataFrame(rows)
    return df_out[[c for c in cols_order if c in df_out.columns]]


def _upload_excel_to_db(uploaded_file) -> tuple:
    """업로드 엑셀 → DB 일괄 삽입. (성공건수, 실패목록) 반환"""
    try:
        df_up = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        return 0, [f"파일 읽기 실패: {e}"]

    success, errors = 0, []
    for idx, row in df_up.iterrows():
        try:
            record = {}
            for field_col in _FIELD_COLS:
                val = row.get(field_col, "")
                record[field_col] = "" if pd.isna(val) else str(val).strip()
            insert_accident(record, source_file="엑셀 일괄업로드")
            success += 1
        except Exception as e:
            errors.append(f"행 {idx+2}: {e}")
    return success, errors


with tab_data:
    st.subheader("📋 사고 데이터 조회 / 엑셀 관리")
    st.caption("DB에 저장된 사고 데이터를 43개 전체 항목으로 조회하고, 엑셀로 다운로드·업로드할 수 있습니다.")

    df_mgmt = get_all_accidents()

    # ── KPI ────────────────────────────────────────────────────
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("전체 레코드",   f"{len(df_mgmt)}건")
    mc2.metric("전체 항목 수",  "43개 필드")
    if not df_mgmt.empty and "risk_grade" in df_mgmt.columns:
        mc3.metric("High 이상",
                   f"{int((df_mgmt['risk_grade'].isin(['High','Critical'])).sum())}건")
    if not df_mgmt.empty and "발생일자" in df_mgmt.columns:
        latest = df_mgmt["발생일자"].dropna().max()
        mc4.metric("최근 발생일", str(latest)[:10] if latest else "-")
    st.divider()

    # ── [A] 조회 필터 + 테이블 ─────────────────────────────────
    st.markdown("#### 🔍 데이터 조회")

    with st.expander("필터 조건", expanded=False):
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            grade_opts = ["전체"] + ["Critical", "High", "Medium", "Low"]
            sel_grade  = st.selectbox("위험 등급", grade_opts, key="mgmt_grade")
        with fc2:
            rail_opts2 = ["전체"] + (
                sorted(df_mgmt["철도구분"].dropna().unique().tolist())
                if not df_mgmt.empty else []
            )
            sel_rail2 = st.selectbox("철도구분", rail_opts2, key="mgmt_rail")
        with fc3:
            line_opts2 = ["전체"] + (
                sorted(df_mgmt["노선"].dropna().unique().tolist())
                if not df_mgmt.empty else []
            )
            sel_line2 = st.selectbox("노선", line_opts2, key="mgmt_line")
        with fc4:
            yr_opts2 = ["전체"] + (
                sorted(df_mgmt["발생일자"].dropna().str[:4].unique().tolist(),
                       reverse=True)
                if not df_mgmt.empty else []
            )
            sel_yr2 = st.selectbox("연도", yr_opts2, key="mgmt_yr")

    df_view = df_mgmt.copy()
    if sel_grade != "전체" and "risk_grade" in df_view.columns:
        df_view = df_view[df_view["risk_grade"] == sel_grade]
    if sel_rail2 != "전체":
        df_view = df_view[df_view["철도구분"] == sel_rail2]
    if sel_line2 != "전체":
        df_view = df_view[df_view["노선"] == sel_line2]
    if sel_yr2 != "전체":
        df_view = df_view[df_view["발생일자"].str.startswith(sel_yr2)]

    st.caption(f"조회 결과: {len(df_view)}건")

    if df_view.empty:
        st.info("조건에 해당하는 데이터가 없습니다.")
    else:
        # 표시할 컬럼 선택기
        show_all = st.toggle("43개 전체 항목 표시", value=False)
        if show_all:
            disp_df = _build_download_df(df_view)
        else:
            quick_cols = [c for c in [
                "id", "발생일자", "등록기관", "노선", "이벤트소분류",
                "근본원인그룹", "사망자수", "부상자수",
                "피해액_백만원", "최대지연시간_분",
                "risk_grade", "risk_score",
            ] if c in df_view.columns]
            disp_df = df_view[quick_cols].copy()

        st.dataframe(
            disp_df,
            use_container_width=True,
            hide_index=True,
            height=360,
            column_config={
                "risk_score": st.column_config.ProgressColumn(
                    "위험점수", format="%.0f", min_value=0, max_value=100
                )
            } if "risk_score" in disp_df.columns else {},
        )

    st.divider()

    # ── [B] 엑셀 다운로드 ──────────────────────────────────────
    st.markdown("#### ⬇️ 엑셀 일괄 다운로드")
    st.caption("현재 필터 조건에 해당하는 데이터를 43개 전체 항목으로 다운로드합니다.")

    dl_col1, dl_col2 = st.columns([2, 1])
    with dl_col1:
        dl_scope = st.radio(
            "다운로드 범위",
            ["현재 필터 결과", "전체 데이터"],
            horizontal=True, key="dl_scope",
        )
    with dl_col2:
        include_meta = st.checkbox("부가 컬럼 포함 (id, 위험점수, 등급)", value=True)

    if st.button("📊 엑셀 파일 생성", type="primary", key="dl_btn"):
        df_dl = df_view if dl_scope == "현재 필터 결과" else df_mgmt
        if df_dl.empty:
            st.warning("다운로드할 데이터가 없습니다.")
        else:
            with st.spinner("엑셀 생성 중..."):
                dl_df = _build_download_df(df_dl)
                if not include_meta:
                    dl_df = dl_df[[c for c in _FIELD_COLS if c in dl_df.columns]]
                xlsx_bytes = _df_to_excel_bytes(dl_df)

            fname_dl = f"사고데이터_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label=f"⬇️ 다운로드 ({len(dl_df)}건 / {len(dl_df.columns)}개 컬럼)",
                data=xlsx_bytes,
                file_name=fname_dl,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_final",
            )
            st.success(f"✅ {len(dl_df)}건, {len(dl_df.columns)}개 항목 엑셀 준비 완료")

    st.divider()

    # ── [C] 엑셀 업로드 ────────────────────────────────────────
    st.markdown("#### ⬆️ 엑셀 일괄 업로드")

    with st.expander("📌 업로드 형식 안내", expanded=False):
        st.markdown("""
**업로드 엑셀 규칙**
- 1행: 헤더 (아래 43개 필드명과 동일해야 함)
- 2행~: 데이터 (빈 셀은 공백 처리됨)
- 날짜: `YYYY-MM-DD`, 시간: `HH:MM`, 숫자 필드: 숫자만

**43개 필드명 목록**
        """)
        col_names_df = pd.DataFrame(
            [{"번호": i+1, "필드명": n, "설명": d} for i, (n, d) in enumerate(COLUMNS)]
        )
        st.dataframe(col_names_df, hide_index=True, use_container_width=True, height=300)

        # 업로드용 빈 양식 다운로드
        empty_df  = pd.DataFrame(columns=_FIELD_COLS)
        tpl_bytes = _df_to_excel_bytes(empty_df)
        st.download_button(
            "📥 빈 양식 다운로드 (템플릿)",
            data=tpl_bytes,
            file_name="사고데이터_입력양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="tpl_dl",
        )

    ul_file = st.file_uploader(
        "엑셀 파일 업로드 (.xlsx)",
        type=["xlsx"],
        key="ul_file",
    )

    if ul_file is not None:
        try:
            df_preview = pd.read_excel(ul_file, engine="openpyxl", nrows=5)
            st.markdown(f"**미리보기** (상위 5행 / 전체 컬럼: {df_preview.shape[1]}개)")
            st.dataframe(df_preview, use_container_width=True, hide_index=True)

            # 필드 매칭 검사
            matched   = [c for c in _FIELD_COLS if c in df_preview.columns]
            unmatched = [c for c in _FIELD_COLS if c not in df_preview.columns]
            st.caption(
                f"✅ 매칭된 필드: {len(matched)}개 / "
                f"⬜ 누락 필드: {len(unmatched)}개"
                + (f" — {unmatched[:5]}" if unmatched else "")
            )
        except Exception as e:
            st.error(f"파일 미리보기 실패: {e}")
            ul_file = None

    if ul_file is not None:
        if st.button("🚀 DB에 일괄 저장", type="primary", key="ul_btn"):
            ul_file.seek(0)
            with st.spinner("업로드 처리 중..."):
                ok_cnt, err_list = _upload_excel_to_db(ul_file)

            if ok_cnt > 0:
                st.success(f"✅ {ok_cnt}건 저장 완료")
                st.rerun()
            if err_list:
                st.warning(f"⚠️ {len(err_list)}건 실패")
                with st.expander("오류 상세"):
                    for e in err_list[:20]:
                        st.text(e)


# ════════════════════════════════════════
# TAB 2. 대시보드
# ════════════════════════════════════════
with tab2:
    st.subheader("📊 사고 데이터 대시보드")
    df_all = get_all_accidents()

    if df_all.empty:
        st.info("데이터 없음. Tab 1에서 보고서를 입력하세요.")
    else:
        trends = analyze_trends(df_all)
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("총 사고",         f"{trends.get('total',0)}건")
        k2.metric("High 위험",       f"{trends.get('high_risk',0)}건")
        k3.metric("평균 위험점수",   f"{trends.get('avg_risk_score',0):.1f}점")
        k4.metric("총 사망자",       f"{trends.get('total_deaths',0)}명")
        k5.metric("총 부상자",       f"{trends.get('total_injured',0)}명")
        st.divider()

        if len(df_all) >= 50:
            df_all = run_anomaly_detection(df_all)
            n_anom = int(df_all['is_anomaly'].sum()) if 'is_anomaly' in df_all.columns else 0
            if n_anom > 0:
                st.warning(f"⚠️ Isolation Forest: {n_anom}건 이상 패턴 탐지")
        else:
            st.caption(f"💡 {len(df_all)}/50건 — 50건 이상 시 이상탐지 활성화")

        r1c1, r1c2 = st.columns(2)
        with r1c1:
            st.markdown("##### 위험 등급 분포")
            if 'risk_grade' in df_all.columns:
                gc = df_all['risk_grade'].value_counts().reset_index(); gc.columns=['등급','건수']
                st.altair_chart(alt.Chart(gc).mark_bar(cornerRadiusTopLeft=4,cornerRadiusTopRight=4).encode(
                    x=alt.X('등급',sort=['High','Medium','Low']), y='건수',
                    color=alt.Color('등급',scale=alt.Scale(domain=['High','Medium','Low'],range=['#e74c3c','#f39c12','#27ae60']),legend=None),
                    tooltip=['등급','건수']).properties(height=250), use_container_width=True)

        with r1c2:
            st.markdown("##### 노선별 사고 건수")
            if '노선' in df_all.columns:
                lc = df_all['노선'].value_counts().head(10).reset_index(); lc.columns=['노선','건수']
                st.altair_chart(alt.Chart(lc).mark_bar(color='#2980b9').encode(
                    x='건수', y=alt.Y('노선',sort='-x'), tooltip=['노선','건수']).properties(height=250),
                    use_container_width=True)

        r2c1, r2c2 = st.columns(2)
        with r2c1:
            st.markdown("##### 근본원인 분포")
            if '근본원인그룹' in df_all.columns:
                cc = df_all['근본원인그룹'].dropna().value_counts().reset_index(); cc.columns=['원인','건수']
                st.altair_chart(alt.Chart(cc).mark_arc(innerRadius=50).encode(
                    theta='건수', color=alt.Color('원인',legend=alt.Legend(orient='right')),
                    tooltip=['원인','건수']).properties(height=250), use_container_width=True)

        with r2c2:
            st.markdown("##### 이벤트 소분류 Top 10")
            if '이벤트소분류' in df_all.columns:
                ec = df_all['이벤트소분류'].dropna().value_counts().head(10).reset_index(); ec.columns=['유형','건수']
                st.altair_chart(alt.Chart(ec).mark_bar(color='#8e44ad').encode(
                    x='건수', y=alt.Y('유형',sort='-x'), tooltip=['유형','건수']).properties(height=250),
                    use_container_width=True)

        if 'risk_score' in df_all.columns and df_all['risk_score'].notna().any():
            st.markdown("##### 위험 점수 분포")
            hd = df_all[['risk_score']].dropna()
            st.altair_chart(alt.Chart(hd).mark_bar(color='#e74c3c',opacity=0.7).encode(
                x=alt.X('risk_score:Q',bin=alt.Bin(maxbins=20),title='위험 점수'),
                y=alt.Y('count()',title='건수'), tooltip=[alt.Tooltip('risk_score:Q',bin=True),'count()']
            ).properties(height=160), use_container_width=True)

        st.divider()
        st.markdown("##### 📋 전체 데이터")
        disp_cols = [c for c in ['id','발생일자','노선','이벤트소분류','근본원인그룹','사망자수','피해액_백만원','risk_grade','risk_score'] if c in df_all.columns]
        st.dataframe(df_all[disp_cols].head(200), use_container_width=True, hide_index=True, height=280,
            column_config={'risk_score':st.column_config.ProgressColumn('위험점수',format='%.0f',min_value=0,max_value=100)})

        with st.expander("🗑️ 데이터 삭제"):
            del_id = st.number_input("삭제할 ID", min_value=1, step=1)
            if st.button("삭제", type="primary"):
                delete_accident(int(del_id)); st.success(f"ID {del_id} 삭제"); st.rerun()

# ════════════════════════════════════════════════════════════
# TAB 3. ⚠️  위험도 평가 매트릭스 (L×C 철도 안전 기준)
# ════════════════════════════════════════════════════════════
# 설계 기준:
#   발생가능성 L(1~5): 이벤트소분류 누적 발생건수 → 5분위 정규화
#   영향도     C(1~5): EFI합(사망+부상/100) + 피해액합 → 직접 5단계 산정
#                      (철도 등가사망 기준: 부상자 100명 = 사망자 1명)
#   위험도 점수 R = L × C  (1~25)
#   Critical 20~25 / High 15~19 / Medium 8~14 / Low 1~7
# ════════════════════════════════════════════════════════════

RISK_GRADE_TABLE = [
    {"grade":"Critical","lo":20,"hi":25,"color":"#C0392B","bg":"#FADBD8","label":"🔴 Critical","action":"운영 중단 검토, 긴급 보수, 일 단위 모니터링","deadline":"즉시"},
    {"grade":"High",    "lo":15,"hi":19,"color":"#E67E22","bg":"#FDEBD0","label":"🟠 High",    "action":"30일 이내 보수 계획 수립, 주 단위 점검",    "deadline":"30일"},
    {"grade":"Medium",  "lo":8, "hi":14,"color":"#F1C40F","bg":"#FEF9E7","label":"🟡 Medium",  "action":"90일 이내 조치, 정기점검 주기 단축",         "deadline":"90일"},
    {"grade":"Low",     "lo":1, "hi":7, "color":"#27AE60","bg":"#EAFAF1","label":"🟢 Low",     "action":"정기점검 유지, 연간 위험도 재평가",           "deadline":"1년"},
]

def _grade_info(r):
    for g in RISK_GRADE_TABLE:
        if g["lo"] <= r <= g["hi"]:
            return g
    return RISK_GRADE_TABLE[-1]

def _freq_to_L(cnt, max_cnt):
    if max_cnt == 0: return 1
    ratio = cnt / max_cnt
    if ratio >= 0.80: return 5
    if ratio >= 0.60: return 4
    if ratio >= 0.40: return 3
    if ratio >= 0.20: return 2
    return 1

def _score_to_C(avg_score):
    """기존 호환용 (단독 avg_score 기반)"""
    if avg_score >= 80: return 5
    if avg_score >= 60: return 4
    if avg_score >= 40: return 3
    if avg_score >= 20: return 2
    return 1

def _C_from_impact(efi_sum: float, damage_sum: float,
                   max_efi: float, max_damage: float,
                   death_sum: float = 0, injury_sum: float = 0,
                   record_count: int = 1) -> int:
    """
    영향도(C) 직접 산정 — 철도 위험도 평가 기준 (상대+절대 이중 기준)

    [상대 기준] 전체 이벤트유형 대비 비율 정규화
      인명점수(0~60) + 물적점수(0~40) → 합산 0~100

    [절대 기준] 철도 인명안전 우선 원칙
      사망자 ≥ 1  → C ≥ 4  (사망 발생 = 중대 이상)
      사망자 ≥ 3  → C = 5  (다수 사망 = 치명)
      부상자 평균 ≥ 20 → C ≥ 4
      부상자 평균 ≥ 5  → C ≥ 3

    최종 C = max(상대기준C, 절대기준C)
    """
    # ── 상대 기준 ─────────────────────────────────────────────
    inj_score = (efi_sum   / max(max_efi,    0.001)) * 60 if max_efi    > 0 else 0
    dmg_score = (damage_sum / max(max_damage, 0.001)) * 40 if max_damage > 0 else 0
    total = min(inj_score + dmg_score, 100)
    c_rel = (5 if total >= 70 else 4 if total >= 50 else 3 if total >= 30 else 2 if total >= 10 else 1)

    # ── 절대 기준 ─────────────────────────────────────────────
    avg_inj = injury_sum / max(record_count, 1)
    c_abs = 1
    if death_sum >= 1:    c_abs = 4   # 사망 발생 → 중대 이상
    if death_sum >= 3:    c_abs = 5   # 다수 사망 → 치명
    if avg_inj  >= 20:    c_abs = max(c_abs, 4)
    if avg_inj  >= 5:     c_abs = max(c_abs, 3)

    return max(c_rel, c_abs)

def _R_grade(r):
    if r >= 20: return "Critical"
    if r >= 15: return "High"
    if r >= 8:  return "Medium"
    return "Low"

with tab3:
    st.subheader("⚠️ 철도 위험도 평가 매트릭스 (L×C)")
    st.caption("철도안전관리체계 기술기준 준용 | 발생가능성(L) × 영향도(C) = 위험도 점수 (1~25)")

    df_rm = get_all_accidents()

    if df_rm.empty:
        st.info("데이터가 없습니다. Tab 1에서 보고서를 입력하세요.")
    else:
        # ── 필터 패널 ────────────────────────────────────────
        with st.expander("🔍 분석 필터", expanded=False):
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                rail_opts = ["전체"] + sorted(df_rm["철도구분"].dropna().unique().tolist())
                sel_rail  = st.selectbox("철도구분", rail_opts, key="rm_rail")
            with fc2:
                line_opts = ["전체"] + sorted(df_rm["노선"].dropna().unique().tolist())
                sel_line  = st.selectbox("노선", line_opts, key="rm_line")
            with fc3:
                yr_list = sorted(df_rm["발생일자"].dropna().str[:4].unique().tolist(), reverse=True)
                yr_opts = ["전체"] + yr_list
                sel_yr  = st.selectbox("연도", yr_opts, key="rm_yr")

        df_f = df_rm.copy()
        if sel_rail != "전체": df_f = df_f[df_f["철도구분"] == sel_rail]
        if sel_line != "전체": df_f = df_f[df_f["노선"] == sel_line]
        if sel_yr   != "전체": df_f = df_f[df_f["발생일자"].str.startswith(sel_yr)]

        if df_f.empty:
            st.warning("선택한 조건에 해당하는 데이터가 없습니다.")
            st.stop()

        # ── L·C·R 산출 ───────────────────────────────────────
        agg = (
            df_f.groupby("이벤트소분류", dropna=True)
            .agg(
                발생건수       =("id",           "count"),
                평균위험점수   =("risk_score",    "mean"),
                사망자합       =("사망자수",       "sum"),
                부상자합       =("부상자수",       "sum"),
                피해액합       =("피해액_백만원",  "sum"),
                최대지연       =("최대지연시간_분","max"),
                근본원인그룹   =("근본원인그룹",   lambda x: x.mode().iloc[0] if (len(x) > 0 and len(x.mode()) > 0) else "-"),
            )
            .reset_index()
        )

        # ── 등가사망지수(EFI) = 사망자 + 부상자 / 100 ────────
        # 철도분야 기준: 부상자 100명 = 사망자 1명
        agg["EFI"] = agg["사망자합"] + agg["부상자합"] / 100.0

        # ── 발생가능성 L(1~5): 발생건수 5분위 ────────────────
        max_cnt    = agg["발생건수"].max()
        agg["L"]   = agg["발생건수"].apply(lambda x: _freq_to_L(x, max_cnt))

        # ── 영향도 C(1~5): EFI합 + 피해액합 + 절대기준 ────────
        # 상대점수(EFI/피해액 비율) + 절대기준(사망자≥1→C≥4) 이중 적용
        max_efi    = agg["EFI"].max()
        max_damage = agg["피해액합"].max()
        agg["C"]   = agg.apply(
            lambda r: _C_from_impact(
                r["EFI"], r["피해액합"], max_efi, max_damage,
                death_sum=r["사망자합"],
                injury_sum=r["부상자합"],
                record_count=max(r["발생건수"], 1),
            ),
            axis=1
        )

        agg["R"]       = agg["L"] * agg["C"]
        agg["위험등급"] = agg["R"].apply(_R_grade)
        agg["평균위험점수"] = agg["평균위험점수"].round(1)
        agg["피해액합"]     = agg["피해액합"].round(1)
        agg["EFI"]          = agg["EFI"].round(2)
        agg = agg.sort_values("R", ascending=False).reset_index(drop=True)

        # ── KPI 요약 ─────────────────────────────────────────
        gc = agg["위험등급"].value_counts()
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("분석 이벤트 유형", f"{len(agg)}종")
        k2.metric("🔴 Critical", f"{gc.get('Critical',0)}종", delta="즉시 대응", delta_color="inverse")
        k3.metric("🟠 High",     f"{gc.get('High',0)}종",     delta="30일 내",  delta_color="inverse")
        k4.metric("🟡 Medium",   f"{gc.get('Medium',0)}종",   delta="90일 내",  delta_color="off")
        k5.metric("🟢 Low",      f"{gc.get('Low',0)}종",      delta="정기점검", delta_color="off")
        st.divider()

        # ════════════════════════════════════════════════════
        # [A] 5×5 리스크 매트릭스
        # ════════════════════════════════════════════════════
        col_mat, col_legend = st.columns([3, 1])

        with col_mat:
            st.markdown("#### 📊 5×5 위험도 평가 매트릭스")

            # 배경 격자
            grid_rows = []
            for lv in range(1, 6):
                for cv in range(1, 6):
                    rv = lv * cv
                    gi = _grade_info(rv)
                    grid_rows.append({"L": lv, "C": cv, "R": rv,
                                      "등급": gi["grade"], "색상": gi["color"]})
            df_grid = pd.DataFrame(grid_rows)

            # 셀 집계 (같은 L,C 묶음)
            cell_rows = []
            for _, row in agg.iterrows():
                cell_rows.append({
                    "L": row["L"], "C": row["C"], "R": row["R"],
                    "이벤트소분류": row["이벤트소분류"],
                    "발생건수": row["발생건수"],
                    "위험등급": row["위험등급"],
                    "사망자": int(row["사망자합"]),
                    "부상자": int(row["부상자합"]),
                })
            df_bubble = pd.DataFrame(cell_rows)

            def _join_types(grp):
                return pd.Series({
                    "유형목록":   " / ".join(grp["이벤트소분류"].tolist()),
                    "유형수":     int(len(grp)),
                    "총발생건수": int(grp["발생건수"].sum()),
                    "총사망자":   int(grp["사망자"].sum()),
                    "총부상자":   int(grp["부상자"].sum()),
                })
            df_cell = (
                df_bubble.groupby(["L","C","위험등급","R"], group_keys=False)
                .apply(_join_types)
                .reset_index()
            )
            df_cell["셀라벨"] = df_cell.apply(
                lambda r: r["유형목록"] if r["유형수"] == 1
                else r["유형목록"].split(" / ")[0] + f" 외 {int(r['유형수'])-1}건",
                axis=1
            )

            # Altair: 배경 격자
            base = alt.Chart(df_grid).encode(
                x=alt.X("C:O", title="영향도 (C) →",
                         axis=alt.Axis(labelAngle=0, values=[1,2,3,4,5])),
                y=alt.Y("L:O", title="발생가능성 (L) ↑", sort="descending"),
            )
            bg_layer = base.mark_rect(stroke="white", strokeWidth=2).encode(
                color=alt.Color("색상:N", scale=None, legend=None),
                tooltip=[
                    alt.Tooltip("L:O", title="발생가능성"),
                    alt.Tooltip("C:O", title="영향도"),
                    alt.Tooltip("R:Q", title="위험도 점수"),
                    alt.Tooltip("등급:N", title="등급"),
                ]
            )
            score_layer = base.mark_text(
                align="right", baseline="top", dx=22, dy=-22,
                size=12, opacity=0.6, fontWeight="bold", color="white"
            ).encode(text=alt.Text("R:Q"))

            # Altair: 버블
            bbase = alt.Chart(df_cell).encode(
                x=alt.X("C:O"),
                y=alt.Y("L:O", sort="descending"),
            )
            bubble_layer = bbase.mark_circle(
                opacity=0.88, stroke="white", strokeWidth=1.5
            ).encode(
                size=alt.Size("총발생건수:Q",
                    scale=alt.Scale(range=[200, 2000]),
                    legend=alt.Legend(title="발생건수", orient="bottom")),
                color=alt.value("#1A1A2E"),
                tooltip=[
                    alt.Tooltip("유형목록:N",   title="이벤트 유형"),
                    alt.Tooltip("유형수:Q",     title="유형 수"),
                    alt.Tooltip("총발생건수:Q", title="총 발생건수"),
                    alt.Tooltip("R:Q",          title="위험도 점수"),
                    alt.Tooltip("위험등급:N",   title="등급"),
                    alt.Tooltip("총사망자:Q",   title="사망자"),
                    alt.Tooltip("총부상자:Q",   title="부상자"),
                ]
            )
            bubble_txt = bbase.mark_text(
                size=10, color="white", fontWeight="bold", dy=1
            ).encode(text=alt.Text("유형수:Q"))

            matrix_chart = (
                alt.layer(bg_layer, score_layer, bubble_layer, bubble_txt)
                .properties(width=500, height=420)
                .configure_axis(labelFontSize=11, titleFontSize=12)
            )
            st.altair_chart(matrix_chart, use_container_width=True)
            st.caption("● 버블 크기 = 발생건수 | 버블 내 숫자 = 해당 셀 이벤트 유형 수 | 마우스 오버로 상세 확인")

            # 범례 보조 표
            leg_data = pd.DataFrame([
                {"등급":"Critical","점수 범위":"20~25","색상":"🔴","대응":"즉시"},
                {"등급":"High",    "점수 범위":"15~19","색상":"🟠","대응":"30일"},
                {"등급":"Medium",  "점수 범위":"8~14", "색상":"🟡","대응":"90일"},
                {"등급":"Low",     "점수 범위":"1~7",  "색상":"🟢","대응":"1년"},
            ])
            st.dataframe(leg_data, hide_index=True, use_container_width=True, height=175)

        with col_legend:
            st.markdown("#### 📌 등급 기준")
            for g in RISK_GRADE_TABLE:
                st.markdown(
                    "<div style='background:{bg};border-left:5px solid {color};"
                    "padding:8px 10px;margin-bottom:6px;border-radius:4px;'>"
                    "<b style='color:{color}'>{label}</b><br>"
                    "<span style='font-size:12px'>점수: {lo}~{hi}</span><br>"
                    "<span style='font-size:11px;color:#555'>{action}</span>"
                    "</div>".format(**g),
                    unsafe_allow_html=True
                )
            st.divider()
            st.markdown("#### 🔢 산출 기준")
            st.markdown(
                "<div style='font-size:11px;line-height:1.8'>"
                "<b>L (발생가능성)</b><br>"
                "최다 빈도 대비 비율 → 5분위<br>"
                "≥80%→5 / ≥60%→4 / ≥40%→3<br>"
                "≥20%→2 / &lt;20%→1<br><br>"
                "<b>C (영향도) — 피해 직접 산정</b><br>"
                "인명(EFI)점수 + 물적점수<br>"
                "EFI = 사망 + 부상÷100<br>"
                "합산 ≥70→5 / ≥50→4<br>"
                "≥30→3 / ≥10→2 / &lt;10→1"
                "</div>",
                unsafe_allow_html=True
            )

        st.divider()

        # ════════════════════════════════════════════════════
        # [B] 이벤트별 위험도 평가 상세 테이블
        # ════════════════════════════════════════════════════
        st.markdown("#### 📋 이벤트 유형별 위험도 평가 결과")

        grade_filter = st.multiselect(
            "등급 필터",
            options=["Critical","High","Medium","Low"],
            default=["Critical","High","Medium","Low"],
            key="rm_grade_filter"
        )
        df_show = agg[agg["위험등급"].isin(grade_filter)].copy()

        GRADE_BG   = {"Critical":"#FADBD8","High":"#FDEBD0","Medium":"#FEF9E7","Low":"#EAFAF1"}
        GRADE_TEXT = {"Critical":"#922B21","High":"#784212","Medium":"#7D6608","Low":"#1E8449"}

        rename_map = {
            "이벤트소분류":"이벤트 유형",
            "발생건수":"발생건수",
            "L":"발생가능성(L)",
            "C":"영향도(C)",
            "R":"위험도(R=L×C)",
            "위험등급":"위험 등급",
            "EFI":"등가사망(EFI)",
            "사망자합":"사망자(명)",
            "부상자합":"부상자(명)",
            "피해액합":"피해액(백만원)",
            "평균위험점수":"평균 위험점수",
            "근본원인그룹":"주요 원인",
        }
        df_disp = df_show[list(rename_map.keys())].rename(columns=rename_map).copy()
        df_disp["사망자(명)"] = df_disp["사망자(명)"].astype(int)
        df_disp["부상자(명)"] = df_disp["부상자(명)"].astype(int)

        # rename 후 컬럼명 "위험 등급" 기준으로 스타일 적용
        def _style_row(row):
            bg  = GRADE_BG.get(row["위험 등급"], "white")
            clr = GRADE_TEXT.get(row["위험 등급"], "black")
            return [
                "background-color:{bg};color:{clr}".format(bg=bg, clr=clr)
                if c == "위험 등급"
                else "background-color:{bg}".format(bg=bg)
                for c in row.index
            ]

        styled = (
            df_disp.style
            .apply(_style_row, axis=1)
            .format({
                "위험도(R=L×C)":  "{:.0f}",
                "등가사망(EFI)":  "{:.2f}",
                "평균 위험점수":  "{:.1f}",
                "피해액(백만원)": "{:.0f}",
            })
            .bar(subset=["위험도(R=L×C)"], color="#E8D5D5", vmin=0, vmax=25)
            .bar(subset=["발생건수"],       color="#D5E8D5", vmin=0)
        )
        st.dataframe(styled, use_container_width=True, height=420, hide_index=True)

        csv_bytes = df_disp.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button(
            "📥 평가 결과 CSV 다운로드", csv_bytes,
            file_name="위험도평가결과_{}.csv".format(datetime.now().strftime("%Y%m%d")),
            mime="text/csv", key="rm_csv"
        )
        st.divider()

        # ════════════════════════════════════════════════════
        # [C] Critical / High 긴급 대응 카드
        # ════════════════════════════════════════════════════
        urgent = df_show[df_show["위험등급"].isin(["Critical","High"])].head(6)
        if not urgent.empty:
            st.markdown("#### 🚨 긴급 대응 필요 항목")
            for _, row in urgent.iterrows():
                ginfo = _grade_info(row["R"])
                col_a, col_b = st.columns([1, 3])
                with col_a:
                    st.markdown(
                        "<div style='background:{bg};border:2px solid {color};"
                        "border-radius:8px;padding:14px;text-align:center;'>"
                        "<div style='font-size:32px;font-weight:bold;color:{color}'>{R}</div>"
                        "<div style='font-size:13px;color:{color}'>{grade}</div>"
                        "<hr style='margin:6px 0;border-color:{color}'>"
                        "<div style='font-size:11px'>L={L} × C={C}</div>"
                        "<div style='font-size:10px;color:#888'>발생 {cnt}건</div>"
                        "</div>".format(
                            bg=ginfo["bg"], color=ginfo["color"],
                            R=int(row["R"]), grade=ginfo["grade"],
                            L=int(row["L"]), C=int(row["C"]),
                            cnt=int(row["발생건수"])
                        ),
                        unsafe_allow_html=True
                    )
                with col_b:
                    st.markdown(
                        "<div style='background:{bg};border-left:4px solid {color};"
                        "border-radius:0 8px 8px 0;padding:12px 16px;'>"
                        "<b style='font-size:16px;color:{color}'>{evtsub}</b>"
                        "<table style='font-size:12px;width:100%;margin-top:8px;border-collapse:collapse'>"
                        "<tr><td style='width:110px;color:#666'>주요 원인</td><td>{cause}</td></tr>"
                        "<tr><td style='color:#666'>피해 현황</td>"
                        "<td>사망 {dead}명 · 부상 {inj}명 · 피해액 {dmg:.0f}백만원</td></tr>"
                        "<tr><td style='color:#666'>대응 기한</td>"
                        "<td><b style='color:{color}'>{deadline}</b></td></tr>"
                        "<tr><td style='color:#666'>조치 내용</td><td>{action}</td></tr>"
                        "</table></div>".format(
                            bg=ginfo["bg"], color=ginfo["color"],
                            evtsub=row["이벤트소분류"],
                            cause=row["근본원인그룹"],
                            dead=int(row["사망자합"]), inj=int(row["부상자합"]),
                            dmg=row["피해액합"],
                            deadline=ginfo["deadline"], action=ginfo["action"]
                        ),
                        unsafe_allow_html=True
                    )
                st.markdown("")

        st.divider()

        # ════════════════════════════════════════════════════
        # [D] 연도별 위험도 추이
        # ════════════════════════════════════════════════════
        st.markdown("#### 📈 연도별 위험 이벤트 발생 추이 (Top 6 유형)")
        df_trend = df_f.copy()
        df_trend["연도"] = df_trend["발생일자"].str[:4]
        top6 = agg["이벤트소분류"].head(6).tolist()
        trend_agg = (
            df_trend[df_trend["이벤트소분류"].isin(top6)]
            .groupby(["연도","이벤트소분류"])
            .agg(건수=("id","count"), 평균점수=("risk_score","mean"))
            .reset_index()
        )
        trend_agg["평균점수"] = trend_agg["평균점수"].round(1)

        if not trend_agg.empty:
            trend_chart = (
                alt.Chart(trend_agg)
                .mark_line(point=alt.OverlayMarkDef(size=60), strokeWidth=2)
                .encode(
                    x=alt.X("연도:O", title="연도"),
                    y=alt.Y("건수:Q", title="발생건수"),
                    color=alt.Color("이벤트소분류:N",
                        legend=alt.Legend(title="이벤트 유형", orient="right")),
                    tooltip=["연도","이벤트소분류","건수","평균점수"]
                )
                .properties(height=260)
            )
            # 위험도 평균 추이 (오른쪽 축)
            trend_score_chart = (
                alt.Chart(trend_agg)
                .mark_area(opacity=0.12, strokeWidth=1.5)
                .encode(
                    x=alt.X("연도:O"),
                    y=alt.Y("평균점수:Q", title="평균 위험점수"),
                    color=alt.Color("이벤트소분류:N", legend=None),
                )
                .properties(height=260)
            )
            st.altair_chart(
                alt.layer(trend_chart, trend_score_chart).resolve_scale(y="independent"),
                use_container_width=True
            )

        st.divider()

        # ════════════════════════════════════════════════════
        # [E] Bow-Tie — 최우선 위험 구조 분석
        # ════════════════════════════════════════════════════
        if not urgent.empty:
            top_item = urgent.iloc[0]
            st.markdown("#### 🎯 Bow-Tie 위험 구조 — [{}]".format(top_item["이벤트소분류"]))
            st.caption("최우선 위험 항목의 위협(원인) → 핵심이벤트 → 결과 구조와 예방 배리어")

            sample = df_f[df_f["이벤트소분류"] == top_item["이벤트소분류"]].head(5)
            causes = sample["직접원인"].dropna().unique().tolist()[:3]

            results = []
            if int(top_item["사망자합"]) > 0:
                results.append("인명사고 사망 {}명".format(int(top_item["사망자합"])))
            if int(top_item["부상자합"]) > 0:
                results.append("부상 {}명".format(int(top_item["부상자합"])))
            results.append("피해액 {:.0f}백만원".format(top_item["피해액합"]))
            results.append("열차 운행 지연·중단")

            ginfo_top = _grade_info(top_item["R"])

            bt1, bt2, bt3 = st.columns([2, 1, 2])
            with bt1:
                st.markdown("**⚡ 위협 요인 (원인)**")
                for c in (causes if causes else ["원인 데이터 없음"]):
                    st.markdown(
                        "<div style='background:#EBF5FB;border-left:3px solid #2980B9;"
                        "padding:6px 10px;margin:4px 0;border-radius:0 4px 4px 0;font-size:12px'>"
                        "{}</div>".format(c),
                        unsafe_allow_html=True
                    )
                st.markdown(
                    "<div style='background:#F8F9FA;border:1px dashed #AAA;"
                    "padding:6px 10px;margin:4px 0;border-radius:4px;font-size:11px;color:#777'>"
                    "근본원인: {}</div>".format(top_item["근본원인그룹"]),
                    unsafe_allow_html=True
                )
            with bt2:
                st.markdown(
                    "<div style='background:{bg};border:3px solid {color};"
                    "border-radius:50%;width:90px;height:90px;display:flex;flex-direction:column;"
                    "align-items:center;justify-content:center;margin:10px auto;text-align:center;"
                    "font-weight:bold;'>"
                    "<div style='font-size:12px;color:{color}'>{sub}</div>"
                    "<div style='font-size:22px;color:{color}'>{R}점</div>"
                    "<div style='font-size:10px;color:{color}'>{grade}</div>"
                    "</div>".format(
                        bg=ginfo_top["bg"], color=ginfo_top["color"],
                        sub=top_item["이벤트소분류"],
                        R=int(top_item["R"]), grade=ginfo_top["grade"]
                    ),
                    unsafe_allow_html=True
                )
            with bt3:
                st.markdown("**💥 결과 (영향)**")
                for r in results:
                    st.markdown(
                        "<div style='background:#FDEDEC;border-right:3px solid #C0392B;"
                        "padding:6px 10px;margin:4px 0;border-radius:4px 0 0 4px;font-size:12px;"
                        "text-align:right'>{}</div>".format(r),
                        unsafe_allow_html=True
                    )

            # 예방 배리어
            st.markdown("**🛡️ 예방 조치 배리어**")
            BARRIER_MAP = {
                "인적요인":   ["운전원 정기 안전교육 강화","신호 준수 실시간 모니터링","피로도 관리 시스템"],
                "기술적요인": ["정기 정밀검사 주기 단축","IoT 실시간 상태 모니터링","예방정비(PM) 계획 강화"],
                "환경적요인": ["기상조건별 속도 제한 강화","선로변 자연재해 대비 보강","외부 침입 감지 시스템"],
            }
            barriers = BARRIER_MAP.get(top_item["근본원인그룹"],
                                       ["안전 점검 강화","위험 요소 모니터링","비상 대응 훈련"])
            bc1, bc2, bc3 = st.columns(3)
            for col, barrier in zip([bc1, bc2, bc3], barriers):
                with col:
                    st.markdown(
                        "<div style='background:#EAF7EC;border:1.5px solid #27AE60;"
                        "border-radius:6px;padding:8px 10px;text-align:center;font-size:12px'>"
                        "🛡️ {}</div>".format(barrier),
                        unsafe_allow_html=True
                    )


# ════════════════════════════════════════
# TAB 4. ML 시계열 예측 대시보드 (v2)
# ════════════════════════════════════════
with tab4:
    st.subheader("🔮 사고 발생 예측 대시보드 (ML 시계열 분석)")
    st.caption("사고조사보고서 DB를 학습한 머신러닝 모델로 향후 이벤트 유형별 사고 발생 확률을 예측합니다.")

    df_fc = get_all_accidents()
    n_fc  = len(df_fc)

    if n_fc < 10:
        st.info(f"예측 모델 실행을 위해 최소 10건이 필요합니다. (현재 {n_fc}건)")
    else:
        # ── 전처리 ────────────────────────────────────────────
        df_fc = df_fc.copy()
        df_fc["발생일자"] = pd.to_datetime(df_fc["발생일자"], errors="coerce")
        df_fc = df_fc.dropna(subset=["발생일자"])
        df_fc["YM"] = df_fc["발생일자"].dt.to_period("M")

        if len(df_fc) < 10:
            st.info("날짜 파싱 가능한 데이터가 부족합니다.")
        else:
            # ── 설정 패널 (항상 표시) ─────────────────────────
            # expander 내부 위젯은 닫힌 상태에서 변경 시 X축 도메인
            # 불일치가 발생하므로 상단에 항상 노출
            st.markdown("##### ⚙️ 예측 설정")
            cfg1, cfg2, cfg3, cfg4 = st.columns([2, 1.5, 2, 2])
            with cfg1:
                forecast_months = st.slider(
                    "예측 기간 (개월)", 3, 24, 12, key="fc_months"
                )
            with cfg2:
                ci_label = st.selectbox(
                    "신뢰 구간", ["80%", "90%", "95%"], index=1, key="fc_ci"
                )
            with cfg3:
                model_choice = st.selectbox(
                    "예측 모델",
                    ["선형 회귀", "2차 다항 회귀", "이동평균 (3M)", "지수평활"],
                    key="fc_model",
                )
            with cfg4:
                _all_lines_fc = sorted(df_fc["노선"].dropna().unique().tolist())
                filter_lines = st.multiselect(
                    "노선 필터 (비어있으면 전체)",
                    _all_lines_fc, default=[], key="fc_lines",
                )

            ci_z = {"80%": 1.28, "90%": 1.645, "95%": 1.96}[ci_label]
            if filter_lines:
                df_fc = df_fc[df_fc["노선"].isin(filter_lines)]

            # 이벤트 유형 선택
            if "이벤트소분류" in df_fc.columns:
                _top_evt_all = df_fc["이벤트소분류"].value_counts().head(8).index.tolist()
            else:
                _top_evt_all = []
            if _top_evt_all:
                # ★ 탭 전환 후 session_state에 남은 이전 선택값이 현재 options에
                #   없으면 StreamlitAPIException 발생 → 사전에 정리
                if "fc_events" in st.session_state:
                    _valid = [e for e in st.session_state["fc_events"] if e in _top_evt_all]
                    if not _valid:
                        _valid = _top_evt_all[:min(5, len(_top_evt_all))]
                    st.session_state["fc_events"] = _valid
                sel_events = st.multiselect(
                    "📊 분석 이벤트 유형 선택 (최대 8종)",
                    _top_evt_all,
                    default=_top_evt_all[:min(5, len(_top_evt_all))],
                    key="fc_events",
                )
            else:
                sel_events = []

            st.divider()

            # ── 상수 ──────────────────────────────────────────
            GRADES_FC  = ["Critical", "High", "Medium", "Low"]
            GRADE_CLRS = {
                "Critical": "#C0392B", "High": "#E67E22",
                "Medium":   "#F1C40F", "Low":  "#27AE60",
            }
            MODEL_DESC = {
                "선형 회귀":       "과거의 선형 추세를 외삽합니다. 단순하고 해석이 용이합니다.",
                "2차 다항 회귀":   "가속·감속 곡선 추세를 포착합니다. 선형보다 유연하나 과적합 주의.",
                "이동평균 (3M)":   "최근 3개월 평균을 기준으로 예측합니다. 단기 변동에 강건합니다.",
                "지수평활":        "최근 값에 높은 가중치를 부여합니다. 수준 변화에 빠르게 반응합니다.",
            }

            # ── 월별 집계 ──────────────────────────────────────
            monthly = (
                df_fc.groupby("YM")
                .agg(
                    건수         = ("id",        "count"),
                    평균위험점수 = ("risk_score", "mean"),
                    Critical     = ("risk_grade", lambda x: (x == "Critical").sum()),
                    High         = ("risk_grade", lambda x: (x == "High").sum()),
                    Medium       = ("risk_grade", lambda x: (x == "Medium").sum()),
                    Low          = ("risk_grade", lambda x: (x == "Low").sum()),
                )
                .reset_index()
                .sort_values("YM")
            )
            monthly["YM_str"]       = monthly["YM"].astype(str)
            monthly["평균위험점수"] = monthly["평균위험점수"].fillna(0)

            hist_ym    = monthly["YM_str"].tolist()
            last_ym    = monthly["YM"].iloc[-1]
            future_ym  = pd.period_range(start=last_ym + 1, periods=forecast_months, freq="M")
            future_str = [str(p) for p in future_ym]
            # ★ 전체 X 도메인을 미리 확정 → 모든 차트에 공유
            all_ym_domain = hist_ym + future_str

            # ── 예측 모델 함수 ─────────────────────────────────
            def _do_forecast(y: np.ndarray, steps: int, z: float, method: str):
                y = y.astype(float)
                if len(y) == 0:
                    z0 = np.zeros(steps)
                    return z0, z0.copy(), z0.copy(), 0.0, 0.0

                if method == "선형 회귀":
                    try:
                        from sklearn.linear_model import LinearRegression
                        t  = np.arange(len(y)).reshape(-1, 1)
                        lr = LinearRegression().fit(t, y)
                        resid = y - lr.predict(t)
                        std   = float(np.std(resid))
                        t_f   = np.arange(len(y), len(y) + steps).reshape(-1, 1)
                        fc    = np.maximum(lr.predict(t_f), 0)
                        return (fc, np.maximum(fc - z*std, 0), fc + z*std,
                                float(lr.coef_[0]), float(lr.score(t, y)))
                    except Exception:
                        pass

                elif method == "2차 다항 회귀":
                    try:
                        from sklearn.preprocessing import PolynomialFeatures
                        from sklearn.linear_model import LinearRegression
                        from sklearn.pipeline import make_pipeline
                        t  = np.arange(len(y)).reshape(-1, 1)
                        md = make_pipeline(PolynomialFeatures(2), LinearRegression())
                        md.fit(t, y)
                        ph  = md.predict(t)
                        std = float(np.std(y - ph))
                        t_f = np.arange(len(y), len(y) + steps).reshape(-1, 1)
                        fc  = np.maximum(md.predict(t_f), 0)
                        ss_res = float(np.sum((y - ph)**2))
                        ss_tot = float(np.sum((y - np.mean(y))**2))
                        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
                        sl = float(md.predict([[len(y)]]) - md.predict([[len(y)-1]])) if len(y) > 1 else 0.0
                        return fc, np.maximum(fc - z*std, 0), fc + z*std, sl, r2
                    except Exception:
                        pass

                elif method == "이동평균 (3M)":
                    w    = min(3, len(y))
                    base = float(np.mean(y[-w:]))
                    std  = float(np.std(y[-w:])) if w > 1 else 0.0
                    fc   = np.full(steps, max(base, 0))
                    return fc, np.maximum(fc - z*std, 0), fc + z*std, 0.0, 0.0

                elif method == "지수평활":
                    alpha = 0.3
                    s     = float(y[0])
                    sm    = [s]
                    for v in y[1:]:
                        s = alpha * float(v) + (1 - alpha) * s
                        sm.append(s)
                    fc  = np.full(steps, max(s, 0))
                    std = float(np.std(y - np.array(sm)))
                    return fc, np.maximum(fc - z*std, 0), fc + z*std, 0.0, 0.0

                # 폴백: 평균
                avg = float(np.mean(y))
                std = float(np.std(y)) if len(y) > 1 else 0.0
                fc  = np.full(steps, max(avg, 0))
                return fc, np.maximum(fc - z*std, 0), fc + z*std, 0.0, 0.0

            # ── 전체 발생건수 예측 ─────────────────────────────
            y_cnt   = monthly["건수"].values.astype(float)
            y_score = monthly["평균위험점수"].values.astype(float)
            fc_cnt,   lo_cnt,   hi_cnt,   slope_cnt,   r2_cnt   = _do_forecast(y_cnt,   forecast_months, ci_z, model_choice)
            fc_score, _,        _,        _,           _        = _do_forecast(y_score, forecast_months, ci_z, model_choice)
            fc_score = np.clip(fc_score, 0, 100)

            # ── 등급 비율 예측 ─────────────────────────────────
            grade_fc = {}
            for g in GRADES_FC:
                y_g = (monthly[g] / monthly["건수"].replace(0, 1)).values.astype(float)
                fg, lg, hg, _, _ = _do_forecast(y_g, forecast_months, ci_z, model_choice)
                grade_fc[g] = {
                    "fc": np.clip(fg, 0, 1),
                    "lo": np.clip(lg, 0, 1),
                    "hi": np.clip(hg, 0, 1),
                }
            for i in range(forecast_months):
                tot = sum(grade_fc[g]["fc"][i] for g in GRADES_FC)
                if tot > 0:
                    for g in GRADES_FC:
                        grade_fc[g]["fc"][i] /= tot

            # ── RF 교차검증 정확도 ─────────────────────────────
            rf_accuracy = None
            if n_fc >= 50:
                try:
                    from sklearn.ensemble import RandomForestClassifier
                    from sklearn.model_selection import cross_val_score
                    _feats = ["사망자수", "부상자수", "피해액_백만원", "최대지연시간_분"]
                    _X = df_fc[_feats].copy()
                    for _c in _feats:
                        _X[_c] = pd.to_numeric(_X[_c], errors="coerce").fillna(0)
                    _y    = df_fc["risk_grade"].fillna("Low")
                    _mask = _y.notna() & _X.notna().all(axis=1)
                    if _mask.sum() >= 20:
                        _clf    = RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1)
                        _scores = cross_val_score(
                            _clf, _X[_mask], _y[_mask],
                            cv=min(5, _mask.sum() // 5), scoring="accuracy",
                        )
                        rf_accuracy = float(_scores.mean())
                except Exception:
                    pass

            # ══════════════════════════════════════════════════
            # KPI 카드
            # ══════════════════════════════════════════════════
            k1, k2, k3, k4, k5 = st.columns(5)
            k1.markdown(
                "<div style='font-size:15px;color:#888;text-align:center'>학습 데이터</div>"
                f"<div style='font-size:20px;font-weight:bold;text-align:center;white-space:nowrap'>{n_fc}건</div>",
                unsafe_allow_html=True
            )
            k2.markdown(
                "<div style='font-size:15px;color:#888;text-align:center'>시계열 범위</div>"
                f"<div style='font-size:20px;font-weight:bold;text-align:center;white-space:nowrap'>{hist_ym[0]} ~ {hist_ym[-1]}</div>",
                unsafe_allow_html=True
            )
            k3.markdown(
                "<div style='font-size:15px;color:#888;text-align:center'>월평균 발생건수</div>"
                f"<div style='font-size:20px;font-weight:bold;text-align:center;white-space:nowrap'>{y_cnt.mean():.1f}건</div>",
                unsafe_allow_html=True
            )
            _trend_val = f"{'▲' if slope_cnt > 0 else '▼'} {abs(slope_cnt):.2f}건/월"
            _trend_color = "#e74c3c" if slope_cnt > 0 else "#27ae60"
            k4.markdown(
                "<div style='font-size:15px;color:#888;text-align:center'>발생 추세</div>"
                f"<div style='font-size:20px;font-weight:bold;text-align:center;white-space:nowrap;color:{_trend_color}'>{_trend_val}</div>",
                unsafe_allow_html=True
            )
            _acc_val = f"{rf_accuracy*100:.0f}%" if rf_accuracy else "N/A (50건 미만)"
            k5.markdown(
                "<div style='font-size:15px;color:#888;text-align:center'>RF 분류 정확도</div>"
                f"<div style='font-size:20px;font-weight:bold;text-align:center;white-space:nowrap'>{_acc_val}</div>",
                unsafe_allow_html=True
            )
            st.divider()

            # ══════════════════════════════════════════════════
            # [A] 통합 차트: 이벤트 유형별 실제+예측+신뢰구간
            # ──────────────────────────────────────────────────
            # ★ 핵심 수정: hist/pred를 별도 DataFrame으로 나누지 않고
            #   all_ym_domain을 X 도메인으로 명시적으로 공유 →
            #   설정 변경 시에도 실선이 사라지지 않음
            # ══════════════════════════════════════════════════
            st.markdown("#### 📈 이벤트 유형별 발생 추이 및 예측 (신뢰구간 포함)")

            # ★ 동일 encoding 객체를 여러 Chart.encode()에 공유하면
            #   Altair 내부에서 dict 직렬화 시 상태가 오염될 수 있음
            #   → 매 호출마다 새 객체를 반환하는 팩토리 함수로 처리
            def _mk_x() -> alt.X:
                return alt.X(
                    "YM:O", title="연월",
                    axis=alt.Axis(labelAngle=-45),
                    scale=alt.Scale(domain=all_ym_domain),
                )

            # ── 통합 long-format 데이터 구성 ──────────────────
            _all_rows: list = []

            # 전체 발생건수 (▶ 전체)
            for j, ym in enumerate(hist_ym):
                _all_rows.append({
                    "YM": ym, "유형": "▶ 전체",
                    "건수": float(y_cnt[j]),
                    "하한": float(y_cnt[j]), "상한": float(y_cnt[j]),
                    "구분": "실제",
                })
            for j, ym in enumerate(future_str):
                _all_rows.append({
                    "YM": ym, "유형": "▶ 전체",
                    "건수": float(fc_cnt[j]),
                    "하한": float(lo_cnt[j]), "상한": float(hi_cnt[j]),
                    "구분": "예측",
                })

            # 이벤트 유형별
            _evt_slope: dict = {}
            for evt in (sel_events or []):
                _y_e = (
                    df_fc[df_fc["이벤트소분류"] == evt]
                    .groupby("YM").size()
                    .reindex(monthly["YM"], fill_value=0)
                    .values.astype(float)
                )
                _fc_e, _lo_e, _hi_e, _sl_e, _ = _do_forecast(
                    _y_e, forecast_months, ci_z, model_choice
                )
                _evt_slope[evt] = _sl_e
                for j, ym in enumerate(hist_ym):
                    _all_rows.append({
                        "YM": ym, "유형": evt,
                        "건수": float(_y_e[j]),
                        "하한": float(_y_e[j]), "상한": float(_y_e[j]),
                        "구분": "실제",
                    })
                for j, ym in enumerate(future_str):
                    _all_rows.append({
                        "YM": ym, "유형": evt,
                        "건수": float(_fc_e[j]),
                        "하한": float(_lo_e[j]), "상한": float(_hi_e[j]),
                        "구분": "예측",
                    })

            _combined = pd.DataFrame(_all_rows)
            _pred_df  = _combined[_combined["구분"] == "예측"].copy()
            _hist_df  = _combined[_combined["구분"] == "실제"].copy()

            # ── 레이어 구성 ────────────────────────────────────
            # 1) 신뢰구간 음영 (예측 구간, 유형별 색상)
            _ci_band = (
                alt.Chart(_pred_df)
                .mark_area(opacity=0.15)
                .encode(
                    x=_mk_x(),
                    y=alt.Y("하한:Q", title="발생건수"),
                    y2=alt.Y2("상한:Q"),
                    color=alt.Color(
                        "유형:N",
                        legend=alt.Legend(title="이벤트 유형", orient="right"),
                    ),
                )
            )
            # 2) 예측 중심선 (점선)
            _pred_line = (
                alt.Chart(_pred_df)
                .mark_line(strokeDash=[5, 3], strokeWidth=2, opacity=0.9)
                .encode(
                    x=_mk_x(),
                    y=alt.Y("건수:Q"),
                    color=alt.Color("유형:N", legend=None),
                    tooltip=[
                        alt.Tooltip("YM:O",   title="연월"),
                        alt.Tooltip("유형:N", title="이벤트 유형"),
                        alt.Tooltip("건수:Q", format=".1f", title="예측 건수"),
                        alt.Tooltip("하한:Q", format=".1f", title=f"하한({ci_label})"),
                        alt.Tooltip("상한:Q", format=".1f", title=f"상한({ci_label})"),
                    ],
                )
            )
            # 3) 실제 실선
            _hist_line = (
                alt.Chart(_hist_df)
                .mark_line(strokeWidth=2.5)
                .encode(
                    x=_mk_x(),
                    y=alt.Y("건수:Q"),
                    color=alt.Color(
                        "유형:N",
                        legend=alt.Legend(title="이벤트 유형", orient="right"),
                    ),
                    tooltip=[
                        alt.Tooltip("YM:O",   title="연월"),
                        alt.Tooltip("유형:N", title="이벤트 유형"),
                        alt.Tooltip("건수:Q", format=".0f", title="실제 건수"),
                    ],
                )
            )
            # 4) 실제 데이터 포인트
            _hist_pts = (
                alt.Chart(_hist_df)
                .mark_point(size=40, filled=True, opacity=0.8)
                .encode(
                    x=_mk_x(),
                    y="건수:Q",
                    color=alt.Color("유형:N", legend=None),
                )
            )
            # 5) 예측 시작 구분선
            _div_line = (
                alt.Chart(pd.DataFrame({"YM": [future_str[0]]}))
                .mark_rule(strokeDash=[4, 4], color="#666", strokeWidth=1.5)
                .encode(x=alt.X("YM:O", scale=alt.Scale(domain=all_ym_domain)))
            ) if future_str else None

            _layers = [_ci_band, _hist_line, _hist_pts, _pred_line]
            if _div_line:
                _layers.append(_div_line)

            st.altair_chart(
                alt.layer(*_layers)
                .properties(height=380)
                .resolve_scale(color="shared"),
                use_container_width=True,
            )
            st.caption(
                f"실선+점=실제 │ 점선+음영=예측({ci_label} 신뢰구간) │ "
                f"수직점선=예측 시작 │ 모델: **{model_choice}** │ "
                f"전체 트렌드 R²={r2_cnt:.2f}"
            )

            # 이벤트 유형별 예측 요약 카드
            if sel_events:
                _card_cols = st.columns(min(len(sel_events) + 1, 4))
                # 전체 카드
                _total_pred_avg = float(np.mean(fc_cnt))
                with _card_cols[0]:
                    _sl_icon = "📈" if slope_cnt > 0.05 else ("📉" if slope_cnt < -0.05 else "➡️")
                    st.markdown(
                        f"<div style='border:2px solid #2c3e50;border-radius:8px;"
                        f"padding:10px;margin:4px 0;text-align:center;background:#f8f9fa'>"
                        f"<div style='font-size:12px;font-weight:bold;color:#2c3e50'>▶ 전체</div>"
                        f"<div style='font-size:20px;color:#2c3e50;font-weight:bold'>{_total_pred_avg:.1f}"
                        f"<span style='font-size:10px;color:#888'>건/월</span></div>"
                        f"<div style='font-size:12px;color:#555'>{_sl_icon} {'증가' if slope_cnt > 0.05 else ('감소' if slope_cnt < -0.05 else '보합')}</div></div>",
                        unsafe_allow_html=True,
                    )
                for _i, _evt in enumerate(sel_events):
                    _pred_vals = [r["건수"] for r in _all_rows if r["유형"] == _evt and r["구분"] == "예측"]
                    if not _pred_vals:
                        continue
                    _avg_c = float(np.mean(_pred_vals))
                    _sl    = _evt_slope.get(_evt, 0)
                    _ico   = "📈" if _sl > 0.05 else ("📉" if _sl < -0.05 else "➡️")
                    _lbl   = "증가" if _sl > 0.05 else ("감소" if _sl < -0.05 else "보합")
                    with _card_cols[(_i + 1) % 4]:
                        st.markdown(
                            f"<div style='border:1px solid #ddd;border-radius:8px;"
                            f"padding:10px;margin:4px 0;text-align:center'>"
                            f"<div style='font-size:12px;font-weight:bold;color:#444'>{_evt}</div>"
                            f"<div style='font-size:20px;color:#2980b9;font-weight:bold'>{_avg_c:.1f}"
                            f"<span style='font-size:10px;color:#888'>건/월</span></div>"
                            f"<div style='font-size:12px;color:#555'>{_ico} {_lbl}</div></div>",
                            unsafe_allow_html=True,
                        )

            # ══════════════════════════════════════════════════
            # [B] 위험 등급 확률 분포 예측
            # ══════════════════════════════════════════════════
            st.markdown("---")
            st.markdown("#### 🎯 향후 위험 등급 확률 분포 예측")

            _grade_rows: list = []
            for _, _row in monthly.iterrows():
                _tot = max(_row["건수"], 1)
                for g in GRADES_FC:
                    _grade_rows.append({
                        "YM": _row["YM_str"], "등급": g,
                        "비율": _row[g] / _tot, "구분": "실제",
                    })
            for i, ym in enumerate(future_str):
                for g in GRADES_FC:
                    _grade_rows.append({
                        "YM": ym, "등급": g,
                        "비율": float(grade_fc[g]["fc"][i]), "구분": "예측",
                    })

            _all_grade_df = pd.DataFrame(_grade_rows)
            _x_grade = alt.X(
                "YM:O", title="연월",
                axis=alt.Axis(labelAngle=-45),
                scale=alt.Scale(domain=all_ym_domain),
            )

            gcol1, gcol2 = st.columns([2, 1])
            with gcol1:
                _grade_area = (
                    alt.Chart(_all_grade_df)
                    .mark_area(opacity=0.82)
                    .encode(
                        x=_x_grade,
                        y=alt.Y("비율:Q", stack="normalize",
                                axis=alt.Axis(format="%"), title="비율(%)"),
                        color=alt.Color(
                            "등급:N",
                            scale=alt.Scale(
                                domain=GRADES_FC,
                                range=[GRADE_CLRS[g] for g in GRADES_FC],
                            ),
                            legend=alt.Legend(title="위험 등급"),
                        ),
                        order=alt.Order("등급:N"),
                        tooltip=[
                            "YM:O", "등급:N",
                            alt.Tooltip("비율:Q", format=".1%"),
                            "구분:N",
                        ],
                    )
                    .properties(height=260)
                )
                if future_str:
                    _vline_g = (
                        alt.Chart(pd.DataFrame({"YM": [future_str[0]]}))
                        .mark_rule(strokeDash=[4, 4], color="#777", strokeWidth=1.5)
                        .encode(x=alt.X("YM:O", scale=alt.Scale(domain=all_ym_domain)))
                    )
                    st.altair_chart(
                        alt.layer(_grade_area, _vline_g).properties(height=260),
                        use_container_width=True,
                    )
                else:
                    st.altair_chart(_grade_area, use_container_width=True)
                st.caption("점선 왼쪽=실제 │ 점선 오른쪽=예측 구간")

            with gcol2:
                st.markdown("##### 예측 기간 평균 확률")
                _avg_g = {g: float(grade_fc[g]["fc"].mean()) for g in GRADES_FC}
                _tot_g = sum(_avg_g.values())
                if _tot_g > 0:
                    _avg_g = {g: v / _tot_g for g, v in _avg_g.items()}
                for g in GRADES_FC:
                    _pct = _avg_g[g] * 100
                    _c   = GRADE_CLRS[g]
                    st.markdown(
                        f"<div style='background:{_c}22;border-left:4px solid {_c};"
                        f"padding:8px 12px;margin:4px 0;border-radius:0 6px 6px 0;'>"
                        f"<b style='color:{_c}'>{g}</b>"
                        f"<span style='float:right;font-size:18px;font-weight:bold;"
                        f"color:{_c}'>{_pct:.1f}%</span></div>",
                        unsafe_allow_html=True,
                    )
                _high_sum = _avg_g.get("Critical", 0) + _avg_g.get("High", 0)
                if _high_sum > 0.35:
                    st.warning(
                        f"⚠️ 고위험(Critical+High) 예측 비율 {_high_sum*100:.0f}% — "
                        "예방 조치 강화 필요"
                    )

            # ══════════════════════════════════════════════════
            # [C] 예측 요약 테이블
            # ══════════════════════════════════════════════════
            st.markdown("---")
            st.markdown("#### 📋 월별 예측 상세")

            _sum_rows: list = []
            for i, ym in enumerate(future_str):
                _dom = max(GRADES_FC, key=lambda g: grade_fc[g]["fc"][i])
                _sum_rows.append({
                    "예측 연월":              ym,
                    "예측 발생건수":          round(float(fc_cnt[i]),   1),
                    f"하한({ci_label})":      round(float(lo_cnt[i]),   1),
                    f"상한({ci_label})":      round(float(hi_cnt[i]),   1),
                    "예측 위험점수":          round(float(fc_score[i]), 1),
                    "Critical 확률":          f"{grade_fc['Critical']['fc'][i]*100:.1f}%",
                    "High 확률":              f"{grade_fc['High']['fc'][i]*100:.1f}%",
                    "Medium 확률":            f"{grade_fc['Medium']['fc'][i]*100:.1f}%",
                    "Low 확률":               f"{grade_fc['Low']['fc'][i]*100:.1f}%",
                    "주요 예측 등급":         _dom,
                })
            _sum_df = pd.DataFrame(_sum_rows)

            _GBGV = {"Critical": "#FADBD8", "High": "#FDEBD0", "Medium": "#FEF9E7", "Low": "#EAFAF1"}
            def _hl_sum(val):
                return f"background-color: {_GBGV.get(val, 'white')}"

            st.dataframe(
                _sum_df.style.map(_hl_sum, subset=["주요 예측 등급"]),
                use_container_width=True, hide_index=True, height=380,
            )

            _csv_fc = _sum_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "📥 예측 결과 CSV", _csv_fc,
                file_name=f"사고예측_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv", key="fc_csv3",
            )

            # ── 모델 설명 ──────────────────────────────────────
            with st.expander(f"ℹ️ 선택 모델 상세 설명: {model_choice}"):
                st.markdown(f"**{model_choice}**: {MODEL_DESC[model_choice]}")
                st.markdown(f"""
**적용 구조:**
- 월별 발생건수 → {model_choice} → {forecast_months}개월 외삽
- 신뢰구간: 예측값 ± {ci_z} × 잔차 표준편차 ({ci_label} 신뢰수준)
- 위험 등급 비율: 등급별 독립 예측 후 합산 정규화 → 확률 분포
{'- RandomForest 분류 정확도(5-fold CV): ' + f'{rf_accuracy*100:.0f}%' if rf_accuracy else '- RandomForest: 50건 이상 시 활성화'}

| 모델 | 특징 | 적합 상황 |
|---|---|---|
| 선형 회귀 | 직선 추세 외삽, 해석 용이 | 꾸준한 증가·감소 |
| 2차 다항 회귀 | 곡선 추세, 가속·감속 포착 | 비선형 변화 구간 |
| 이동평균 (3M) | 최근 3개월 평균, 변동 완화 | 불규칙 변동이 큰 데이터 |
| 지수평활 | 최근값 가중치 강조 | 최근 수준 변화에 민감 반응 |

> **주의:** 통계적 추세 기반 예측이므로 실제 운영 의사결정 전 전문가 검토가 필요합니다.
                """)



with tab_agent:
    render_agent_tab(model_name)