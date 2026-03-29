"""
ui/tab_data.py — 탭2: 데이터 조회/관리 (엑셀 다운로드·업로드)
"""
import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st

from safety_core import get_all_accidents, insert_accident

# ── DB 컬럼명 ↔ 43개 필드명 매핑 ─────────────────────────────
_DB_TO_FIELD = {
    "발생일자": "발생일자", "발생시간": "발생시간", "등록기관": "등록기관",
    "철도구분": "철도구분", "노선": "노선",
    "이벤트대분류": "이벤트대분류", "이벤트중분류": "이벤트중분류", "이벤트소분류": "이벤트소분류",
    "주원인": "주원인", "근본원인그룹": "근본원인그룹", "근본원인유형": "근본원인유형",
    "근본원인상세": "근본원인상세", "직접원인": "직접원인",
    "운행영향유형": "운행영향유형", "지연여부": "지연여부",
    "지연원인": "지연원인", "지연원인상세": "지연원인상세", "지연열차수": "지연열차수",
    "최대지연시간(분)": "최대지연시간(분)",
    "총피해인원": "총피해인원", "사망자수": "사망자수", "부상자수": "부상자수",
    "피해액(백만원)": "피해액(백만원)",
    "행정구역": "행정구역", "발생역A": "발생역A", "발생역B": "발생역B",
    "장소대분류": "장소대분류", "장소중분류": "장소중분류", "상세위치": "상세위치",
    "기상상태": "기상상태", "온도": "온도", "강우량": "강우량", "적설량": "적설량",
    "대상구분": "대상구분", "열차종류": "열차종류", "선로유형": "선로유형",
    "신호시스템유형": "신호시스템유형",
    "고장부품명": "고장부품명", "고장현상": "고장현상", "고장원인": "고장원인",
    "조치내용": "조치내용", "이벤트개요": "이벤트개요", "데이터출처": "데이터출처",
}
_FIELD_COLS = [v for v in _DB_TO_FIELD.values()]
_EXTRA_COLS = ["id", "risk_score", "risk_grade"]


def _df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """DataFrame → xlsx bytes (openpyxl 스타일 적용)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사고데이터"

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="EEF4FB")
    GRADE_FILLS = {
        "Critical": PatternFill("solid", start_color="FADBD8"),
        "High":     PatternFill("solid", start_color="FDEBD0"),
        "Medium":   PatternFill("solid", start_color="FEF9E7"),
        "Low":      PatternFill("solid", start_color="EAFAF1"),
    }

    headers = list(df.columns)
    ws.row_dimensions[1].height = 24
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        ws.column_dimensions[c.column_letter].width = max(12, min(len(str(h)) * 2, 40))

    grade_col_idx = headers.index("risk_grade") + 1 if "risk_grade" in headers else None
    for ri, row in enumerate(df.itertuples(index=False), 2):
        grade = str(getattr(row, "risk_grade", "")) if grade_col_idx else ""
        row_fill = GRADE_FILLS.get(grade, alt_fill if ri % 2 == 0 else None)
        for ci, val in enumerate(row, 1):
            cell_val = "" if val is None or (isinstance(val, float) and np.isnan(val)) else val
            c = ws.cell(row=ri, column=ci, value=cell_val)
            c.font = Font(name="Arial", size=9); c.alignment = left; c.border = border
            if row_fill:
                c.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _build_download_df(df_raw: pd.DataFrame) -> pd.DataFrame:
    """DB 조회 결과 → 43개 필드 + 부가컬럼 DataFrame"""
    rows = []
    for _, r in df_raw.iterrows():
        row = {}
        for db_col, field_col in _DB_TO_FIELD.items():
            row[field_col] = r.get(db_col, "")
        for ec in _EXTRA_COLS:
            row[ec] = r.get(ec, "")
        rows.append(row)
    cols_order = ["id"] + _FIELD_COLS + ["risk_score", "risk_grade"]
    df_out = pd.DataFrame(rows)
    return df_out[[c for c in cols_order if c in df_out.columns]]


def _upload_excel_to_db(uploaded_file) -> tuple:
    """업로드 엑셀 → DB 일괄 삽입. (성공건수, 실패목록) 반환"""
    try:
        df_up = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        return 0, [f"파일 읽기 실패: {e}"]

    success, errors = 0, []
    for idx, row in df_up.iterrows():
        try:
            record = {
                field_col: ("" if pd.isna(row.get(field_col, "")) else str(row.get(field_col, "")).strip())
                for field_col in _FIELD_COLS
            }
            insert_accident(record, source_file="엑셀 일괄업로드")
            success += 1
        except Exception as e:
            errors.append(f"행 {idx+2}: {e}")
    return success, errors


def render_data_tab(columns):
    """
    탭2 데이터 조회/관리 UI.

    Args:
        columns: [(필드명, 설명), ...] 리스트 (업로드 형식 안내에 사용)
    """
    st.subheader("📋 사고 데이터 조회 / 엑셀 관리")
    st.caption("DB에 저장된 사고 데이터를 43개 전체 항목으로 조회하고, 엑셀로 다운로드·업로드할 수 있습니다.")

    df_mgmt = get_all_accidents()

    # ── KPI ────────────────────────────────────────────────────
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("전체 레코드", f"{len(df_mgmt)}건")
    mc2.metric("전체 항목 수", "43개 필드")
    if not df_mgmt.empty and "risk_grade" in df_mgmt.columns:
        mc3.metric("High 이상", f"{int(df_mgmt['risk_grade'].isin(['High','Critical']).sum())}건")
    if not df_mgmt.empty and "발생일자" in df_mgmt.columns:
        latest = df_mgmt["발생일자"].dropna().max()
        mc4.metric("최근 발생일", str(latest)[:10] if latest else "-")
    st.divider()

    # ── [A] 조회 필터 + 테이블 ─────────────────────────────────
    st.markdown("#### 🔍 데이터 조회")
    with st.expander("필터 조건", expanded=False):
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            sel_grade = st.selectbox("위험 등급", ["전체","Critical","High","Medium","Low"], key="mgmt_grade")
        with fc2:
            rail_opts2 = ["전체"] + (sorted(df_mgmt["철도구분"].dropna().unique().tolist()) if not df_mgmt.empty else [])
            sel_rail2 = st.selectbox("철도구분", rail_opts2, key="mgmt_rail")
        with fc3:
            line_opts2 = ["전체"] + (sorted(df_mgmt["노선"].dropna().unique().tolist()) if not df_mgmt.empty else [])
            sel_line2 = st.selectbox("노선", line_opts2, key="mgmt_line")
        with fc4:
            yr_opts2 = ["전체"] + (sorted(df_mgmt["발생일자"].dropna().str[:4].unique().tolist(), reverse=True) if not df_mgmt.empty else [])
            sel_yr2 = st.selectbox("연도", yr_opts2, key="mgmt_yr")

    df_view = df_mgmt.copy()
    if sel_grade != "전체" and "risk_grade" in df_view.columns:
        df_view = df_view[df_view["risk_grade"] == sel_grade]
    if sel_rail2 != "전체":
        df_view = df_view[df_view["철도구분"] == sel_rail2]
    if sel_line2 != "전체":
        df_view = df_view[df_view["노선"] == sel_line2]
    if sel_yr2 != "전체":
        df_view = df_view[df_view["발생일자"].str.startswith(sel_yr2)]

    st.caption(f"조회 결과: {len(df_view)}건")
    if df_view.empty:
        st.info("조건에 해당하는 데이터가 없습니다.")
    else:
        show_all = st.toggle("43개 전체 항목 표시", value=False)
        if show_all:
            disp_df = _build_download_df(df_view)
        else:
            quick_cols = [c for c in [
                "id","발생일자","등록기관","노선","이벤트소분류",
                "근본원인그룹","사망자수","부상자수",
                "피해액(백만원)","최대지연시간(분)",
                "risk_grade","risk_score",
            ] if c in df_view.columns]
            disp_df = df_view[quick_cols].copy()

        st.dataframe(
            disp_df, use_container_width=True, hide_index=True, height=360,
            column_config={
                "risk_score": st.column_config.ProgressColumn("위험점수", format="%.0f", min_value=0, max_value=100)
            } if "risk_score" in disp_df.columns else {},
        )

    st.divider()

    # ── [B] 엑셀 다운로드 ──────────────────────────────────────
    st.markdown("#### ⬇️ 엑셀 일괄 다운로드")
    st.caption("현재 필터 조건에 해당하는 데이터를 43개 전체 항목으로 다운로드합니다.")
    dl_col1, dl_col2 = st.columns([2, 1])
    with dl_col1:
        dl_scope = st.radio("다운로드 범위", ["현재 필터 결과","전체 데이터"], horizontal=True, key="dl_scope")
    with dl_col2:
        include_meta = st.checkbox("부가 컬럼 포함 (id, 위험점수, 등급)", value=True)

    if st.button("📊 엑셀 파일 생성", type="primary", key="dl_btn"):
        df_dl = df_view if dl_scope == "현재 필터 결과" else df_mgmt
        if df_dl.empty:
            st.warning("다운로드할 데이터가 없습니다.")
        else:
            with st.spinner("엑셀 생성 중..."):
                dl_df = _build_download_df(df_dl)
                if not include_meta:
                    dl_df = dl_df[[c for c in _FIELD_COLS if c in dl_df.columns]]
                xlsx_bytes = _df_to_excel_bytes(dl_df)
            fname_dl = f"사고데이터_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label=f"⬇️ 다운로드 ({len(dl_df)}건 / {len(dl_df.columns)}개 컬럼)",
                data=xlsx_bytes, file_name=fname_dl,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_final",
            )
            st.success(f"✅ {len(dl_df)}건, {len(dl_df.columns)}개 항목 엑셀 준비 완료")

    st.divider()

    # ── [C] 엑셀 업로드 ────────────────────────────────────────
    st.markdown("#### ⬆️ 엑셀 일괄 업로드")
    with st.expander("📌 업로드 형식 안내", expanded=False):
        st.markdown("""
**업로드 엑셀 규칙**
- 1행: 헤더 (아래 43개 필드명과 동일해야 함)
- 2행~: 데이터 (빈 셀은 공백 처리됨)
- 날짜: `YYYY-MM-DD`, 시간: `HH:MM`, 숫자 필드: 숫자만

**43개 필드명 목록**
        """)
        col_names_df = pd.DataFrame(
            [{"번호": i+1, "필드명": n, "설명": d} for i, (n, d) in enumerate(columns)]
        )
        st.dataframe(col_names_df, hide_index=True, use_container_width=True, height=300)
        empty_df  = pd.DataFrame(columns=_FIELD_COLS)
        tpl_bytes = _df_to_excel_bytes(empty_df)
        st.download_button(
            "📥 빈 양식 다운로드 (템플릿)", data=tpl_bytes,
            file_name="사고데이터_입력양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="tpl_dl",
        )

    ul_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"], key="ul_file")
    if ul_file is not None:
        try:
            df_preview = pd.read_excel(ul_file, engine="openpyxl", nrows=5)
            st.markdown(f"**미리보기** (상위 5행 / 전체 컬럼: {df_preview.shape[1]}개)")
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            matched   = [c for c in _FIELD_COLS if c in df_preview.columns]
            unmatched = [c for c in _FIELD_COLS if c not in df_preview.columns]
            st.caption(
                f"✅ 매칭된 필드: {len(matched)}개 / ⬜ 누락 필드: {len(unmatched)}개"
                + (f" — {unmatched[:5]}" if unmatched else "")
            )
        except Exception as e:
            st.error(f"파일 미리보기 실패: {e}")
            ul_file = None

    if ul_file is not None:
        if st.button("🚀 DB에 일괄 저장", type="primary", key="ul_btn"):
            ul_file.seek(0)
            with st.spinner("업로드 처리 중..."):
                ok_cnt, err_list = _upload_excel_to_db(ul_file)
            if ok_cnt > 0:
                st.success(f"✅ {ok_cnt}건 저장 완료")
                st.rerun()
            if err_list:
                st.warning(f"⚠️ {len(err_list)}건 실패")
                with st.expander("오류 상세"):
                    for e in err_list[:20]:
                        st.text(e)
